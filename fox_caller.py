#!/usr/bin/env python3
"""
Fox Caller - Standalone Bulk SIP Calling Tool
==============================================
Reads phone numbers from .xlsx files, authenticates with Telicall API,
and makes SIP calls using multiple accounts in parallel.

Processing Logic (Auto Mode):
  - Fixed pool of workers keeps running
  - Each worker gets an account from the shared pool
  - When account dies, worker gets a NEW account and continues
  - When number is answered, worker gets a NEW account and continues
  - Workers only stop when no more accounts OR no more numbers

Processing Logic (Single Mode):
  - Sequential: one number at a time, one account at a time
  - If no answer, try same number with next account

Usage:
    python3 fox_caller.py [directory_path]

If no directory is specified, it uses the current working directory.
"""

import os
import sys
import re
import json
import time
import uuid
import socket
import struct
import hashlib
import base64
import random
import threading
import wave
import audioop
import requests
import string
import openpyxl
from datetime import datetime
from collections import deque

# ============================================================================
#                            CONFIGURATION
# ============================================================================

API_URL = "https://api.telicall.com"
ACCOUNTS_PASSWORD = "@@@GMAQ@@@"
CALL_DURATION = 3             # seconds to stay in call after answer (auto mode - just detect, then hang up)
JUST_DETECT_ANSWER = True     # If True, hang up immediately after detecting answer (auto mode)
SEQ_CALL_DURATION = 62        # seconds to stay in call in sequential mode then hang up and move on
SEQ_JUST_DETECT = False       # In seq mode: keep call alive for 62s then hang up
PARALLEL_ACCOUNTS = 4         # Use 4 accounts simultaneously per number
RINGING_TIMEOUT = 40          # SIP loop iterations (40 * 0.5s = 20s ringing) WAS 140=70s
MIN_ANSWERED_DURATION = 1     # minimum seconds to count as "answered ok"
INSTANT_BYE_CHECKS = 3        # checks for instant BYE after answer (3 * 0.2s = 0.6s)
MAX_RETRIES_PER_NUMBER = 10   # max account attempts for the same number (auto mode)
SINGLE_MAX_RETRIES = 50       # max account attempts for single mode
MAX_CONCURRENT_CALLS = 300    # max parallel calls in auto mode
INIT_BATCH_SIZE = 20          # how many accounts to init in parallel
CACHE_FILENAME = "fox_accounts_cache.json"  # cache file for ready accounts
NUM_WORKERS = 50              # number of persistent worker threads
SIP_CONNECT_TIMEOUT = 8       # seconds for SIP socket connect (was 30!)
RECV_TIMEOUT = 4              # seconds for SIP REGISTER/INVITE recv (was 10!)
MAX_CALL_TIME = 30            # overall max seconds for a SIP call attempt (new!)
API_TIMEOUT = 5               # seconds for Telicall API start call (was 8!)
CALL_HISTORY_FILE = "fox_call_history.json"  # tracks last call time per number
CALL_COOLDOWN = 70            # seconds before re-calling same number
AUTO_CREATE_ACCOUNTS = True   # auto-create accounts when they run out
ACCOUNT_CREATE_BATCH = 15     # how many accounts to create at once (larger batches for speed)
SKIP_BALANCE_CHECK = True     # Skip balance check - accounts will be tested with real numbers

# ============================================================================
#                             ANSI COLORS
# ============================================================================

class C:
    """ANSI color codes for terminal output."""
    RST = '\033[0m'
    BOLD = '\033[1m'
    DIM = '\033[2m'
    # Foreground
    RED = '\033[91m'
    GREEN = '\033[92m'
    YELLOW = '\033[93m'
    BLUE = '\033[94m'
    MAGENTA = '\033[95m'
    CYAN = '\033[96m'
    WHITE = '\033[97m'
    # Bold foreground
    BRED = '\033[1;91m'
    BGREEN = '\033[1;92m'
    BYELLOW = '\033[1;93m'
    BBLUE = '\033[1;94m'
    BCMAGENTA = '\033[1;95m'
    BCYAN = '\033[1;96m'
    BWHITE = '\033[1;97m'
    # Background
    BG_RED = '\033[41m'
    BG_GREEN = '\033[42m'
    BG_BLUE = '\033[44m'
    BG_CYAN = '\033[46m'
    BG_YELLOW = '\033[43m'


def supports_color():
    """Check if terminal supports color."""
    if os.getenv('NO_COLOR'):
        return False
    if os.getenv('COLORTERM'):
        return True
    if not hasattr(sys.stdout, 'isatty'):
        return False
    if not sys.stdout.isatty():
        return False
    term = os.getenv('TERM', '')
    if 'color' in term or 'xterm' in term or 'screen' in term:
        return True
    return False


COLOR = supports_color()


def clr(color_code, text):
    """Apply color to text if color is supported."""
    if COLOR:
        return f"{color_code}{text}{C.RST}"
    return text


def status_icon(result):
    """Get colored status icon for a call result."""
    icons = {
        'answered_ok':   clr(C.BGREEN, 'V'),
        'answered_short':clr(C.GREEN,  'v'),
        'no_answer':     clr(C.YELLOW, '-'),
        'declined':      clr(C.BYELLOW,'X'),
        'busy':          clr(C.BYELLOW,'X'),
        'not_found':     clr(C.MAGENTA, '?'),
        'failed':        clr(C.BRED,   '!'),
    }
    return icons.get(result, clr(C.RED, '?'))


# ============================================================================
#                             GLOBAL STATE
# ============================================================================

accounts = []                  # list of account dicts WITH tokens
used_emails = set()            # emails that have been consumed (dead accounts)
account_index = 0              # current position in accounts list
account_lock = threading.Lock()

# Shared number queue for auto mode
number_queue = deque()         # numbers waiting to be called
number_lock = threading.Lock()
active_calls = 0               # currently active calls
active_lock = threading.Lock()

# Retry tracking for auto mode
number_attempts = {}           # phone -> attempt count
number_attempts_lock = threading.Lock()
number_last_result = {}        # phone -> last call result string
number_last_result_lock = threading.Lock()

# Track which numbers are fully done (answered or max retries exceeded)
done_numbers = set()
done_lock = threading.Lock()

# Results tracking
answered_phones = []
answered_lock = threading.Lock()

# Live call log for display
call_log = []
call_log_lock = threading.Lock()
MAX_LOG_LINES = 12

# Active call tracking (calls currently in RTP/talking phase)
in_call_count = 0
in_call_lock = threading.Lock()

# Stats
stats = {
    "answered": 0,
    "no_answer": 0,
    "busy": 0,
    "failed": 0,
    "not_found": 0,
    "answered_short": 0,
    "accounts_used": 0,
    "accounts_remaining": 0,
    "total_processed": 0,
    "total_numbers": 0,
    "api_fail": 0,
    "skipped": 0,          # numbers never called (accounts ran out)
    "auto_created": 0,     # accounts auto-created during run
}
stats_lock = threading.Lock()

# ============================================================================
#                          TELICALL API HELPERS
# ============================================================================

def get_headers(token=None, device_id=None):
    """Build Telicall API request headers."""
    if not device_id:
        device_id = ''.join(random.choices('0123456789abcdef', k=16))
    return {
        "host": "api.telicall.com",
        "x-request-id": str(uuid.uuid4()),
        "user-agent": "Dalvik/2.1.0",
        "x-app-version": "1.2.1",
        "x-client-device-id": device_id,
        "x-lang": "en",
        "x-os": "android",
        "x-os-version": "11",
        "x-req-timestamp": str(int(time.time() * 1000)),
        "x-req-signature": "-1",
        "content-type": "application/json",
        "x-token": token or ""
    }


def telicall_init_session(device_id):
    """Call POST /init to get a session token for a device."""
    h = get_headers(device_id=device_id)
    h["x-token"] = ""
    body = {
        "countryCode": "eg",
        "deviceName": "Infinix X698",
        "notificationToken": "",
        "oldToken": "",
        "peerKey": str(random.randint(100, 999)),
        "timeZone": "Africa/Cairo",
        "localizationKey": ""
    }
    try:
        r = requests.post(f"{API_URL}/init", json=body, headers=h, timeout=10)
        if r.status_code == 200 and r.json().get('result', {}).get('token'):
            return r.json()['result']['token']
    except:
        pass
    return None


def telicall_start_call(phone, call_token, call_device_id):
    """Start a call via Telicall API and return SIP credentials."""
    if not phone.startswith('+'):
        phone = '+' + phone
    headers = get_headers(token=call_token, device_id=call_device_id)
    try:
        r = requests.post(
            f"{API_URL}/call/outbound/start",
            json={'to': phone, 'source': 'numpad'},
            headers=headers,
            timeout=API_TIMEOUT
        )
        if r.status_code == 200 and r.json().get('result'):
            sip = r.json()['result'].get('sip', {})
            return {
                'user': sip.get('username'),
                'pass': sip.get('password'),
                'domain': sip.get('domain'),
                'port': sip.get('port', 5060),
                'proto': sip.get('protocol', 'tcp'),
                'from': r.json()['result'].get('from', {}).get('msisdn'),
                'to': r.json()['result'].get('to', {}).get('msisdn'),
                'limit': sip.get('callLimit', 60),
                'balance': sip.get('balanceLimit', 60),
            }
        elif r.status_code == 400:
            err_text = r.text.lower()
            if 'balance' in err_text:
                return 'no_balance'
            return {'error': 'call_400'}
        elif r.status_code == 404:
            return {'error': 'call_404'}
        else:
            return {'error': f'call_{r.status_code}'}
    except Exception as e:
        return None


def get_next_account():
    """Get the next available account (thread-safe). Returns dict or None."""
    global account_index
    with account_lock:
        while account_index < len(accounts):
            acc = accounts[account_index]
            account_index += 1
            email = acc.get('email', '')
            if email not in used_emails and (acc.get('x-token') or acc.get('token')):
                return acc
        return None


def get_next_accounts(count):
    """Get up to 'count' available accounts at once (thread-safe). Returns list of dicts."""
    result = []
    with account_lock:
        while account_index < len(accounts) and len(result) < count:
            acc = accounts[account_index]
            account_index += 1
            email = acc.get('email', '')
            if email not in used_emails and (acc.get('x-token') or acc.get('token')):
                result.append(acc)
    return result


def count_remaining():
    """Count remaining unused accounts."""
    with account_lock:
        remaining = 0
        for i in range(account_index, len(accounts)):
            if accounts[i].get('email', '') not in used_emails and (accounts[i].get('x-token') or accounts[i].get('token')):
                remaining += 1
        return remaining


def mark_used(email, status='used'):
    """Mark an account email as used/dead and update cache."""
    with account_lock:
        used_emails.add(email)
    with stats_lock:
        stats["accounts_used"] += 1
    # Update cache file
    if hasattr(mark_used, 'cache_dir'):
        update_cache_after_call(mark_used.cache_dir, email, status)


def start_call_with_account(phone, acc):
    """Start a call using a specific account. Returns SIP info or error."""
    call_token = acc.get('x-token', '') or acc.get('token', '')
    call_device_id = acc.get('x-client-device-id', '') or acc.get('device_id', '')
    email = acc.get('email', '')

    if not call_token:
        mark_used(email, status='dead')
        return None, email

    result = telicall_start_call(phone, call_token, call_device_id)

    if result == 'no_balance':
        # Account has no balance - truly dead
        mark_used(email, status='dead')
        return result, email
    elif result is None:
        # Network/timeout - don't kill account, might be temporary
        return result, email
    elif isinstance(result, dict) and "error" in result:
        err = result.get('error', '')
        if err == 'call_403':
            # 403 = Forbidden - usually number/region block, NOT account dead
            # Don't kill the account! It might work for other numbers.
            return result, email
        elif err == 'call_404':
            # 404 = Number not found - account is fine
            return result, email
        elif err == 'call_400':
            # 400 = Bad request - account might be fine
            return result, email
        else:
            # Other errors - might be account issue, mark dead
            mark_used(email, status='dead')
            return result, email

    # Success
    return result, email


# ============================================================================
#                     ACCOUNT CACHE SYSTEM
# ============================================================================

def get_cache_path(directory):
    """Get the path to the accounts cache file."""
    return os.path.join(directory, CACHE_FILENAME)


def load_accounts_cache(directory):
    """
    Load cached accounts from the cache file.
    Returns dict with ready/dead/used lists, or None if no cache.
    """
    cache_path = get_cache_path(directory)
    if not os.path.exists(cache_path):
        return None
    try:
        with open(cache_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        return data
    except:
        return None


def save_accounts_cache(directory, ready_accounts, dead_emails_set, used_emails_set):
    """
    Save accounts cache to file.
    Stores ready accounts (with tokens), dead accounts, and used accounts.
    """
    cache_path = get_cache_path(directory)
    cache_data = {
        'last_updated': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'ready': [],
        'dead': [],
        'used': []
    }

    # Save ready accounts (with tokens)
    for acc in ready_accounts:
        email = acc.get('email', '')
        if email in dead_emails_set or email in used_emails_set:
            continue
        cache_data['ready'].append({
            'email': acc.get('email', ''),
            'x-token': acc.get('x-token', ''),
            'x-client-device-id': acc.get('x-client-device-id', ''),
            'status': 'ready'
        })

    # Save dead accounts
    for email in dead_emails_set:
        cache_data['dead'].append({'email': email, 'status': 'dead'})

    # Save used/consumed accounts
    for email in used_emails_set:
        if email not in dead_emails_set:
            cache_data['used'].append({'email': email, 'status': 'used'})

    try:
        with open(cache_path, 'w', encoding='utf-8') as f:
            json.dump(cache_data, f, indent=2, ensure_ascii=False)
    except Exception as e:
        pass


def update_cache_after_call(directory, email, status):
    """
    Update the cache file after a call result.
    status: 'used' (account consumed after answered), 'dead' (account failed)
    """
    cache_path = get_cache_path(directory)
    if not os.path.exists(cache_path):
        return

    try:
        with open(cache_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
    except:
        return

    # Remove from ready list if present
    data['ready'] = [a for a in data['ready'] if a.get('email') != email]

    if status == 'dead':
        dead_emails = [a.get('email') for a in data.get('dead', [])]
        if email not in dead_emails:
            data['dead'].append({'email': email, 'status': 'dead'})
    elif status == 'used':
        used_emails_list = [a.get('email') for a in data.get('used', [])]
        if email not in used_emails_list:
            data['used'].append({'email': email, 'status': 'used'})

    data['last_updated'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    try:
        with open(cache_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
    except:
        pass


def merge_cache_with_dan(cache_data, raw_accounts):
    """
    Merge cached accounts with Dan.json accounts.
    - Cached ready accounts are reused directly (skip /init)
    - Cached dead/used accounts are excluded from init
    - New accounts not in cache need /init
    Returns: (ready_accounts, need_init_accounts, dead_emails_set)
    """
    dead_emails = set()
    used_emails_from_cache = set()
    cached_ready = {}

    if cache_data:
        # Collect dead emails
        for acc in cache_data.get('dead', []):
            dead_emails.add(acc.get('email', ''))

        # Collect used emails (treat same as dead - they're consumed)
        for acc in cache_data.get('used', []):
            used_emails_from_cache.add(acc.get('email', ''))

        # Collect ready accounts with tokens
        for acc in cache_data.get('ready', []):
            email = acc.get('email', '')
            token = acc.get('x-token', '')
            device_id = acc.get('x-client-device-id', '')
            if email and token and device_id:
                cached_ready[email] = acc

    ready = []
    need_init = []

    for acc in raw_accounts:
        email = acc.get('email', '')

        # Skip dead accounts
        if email in dead_emails:
            continue

        # Skip used accounts
        if email in used_emails_from_cache:
            continue

        # Check if already cached and ready
        if email in cached_ready:
            cached = cached_ready[email]
            acc['x-token'] = cached.get('x-token', '')
            acc['x-client-device-id'] = cached.get('x-client-device-id', '')
            ready.append(acc)
        else:
            need_init.append(acc)

    return ready, need_init, dead_emails


# ============================================================================
#                     CALL HISTORY SYSTEM (Cooldown Tracking)
# ============================================================================

def get_call_history_path(directory):
    """Get the path to the call history JSON file."""
    return os.path.join(directory, CALL_HISTORY_FILE)


def load_call_history(directory):
    """
    Load call history from JSON file.
    Returns dict {phone_no_plus: timestamp_float}.
    Phone numbers are stored WITHOUT the + prefix for normalization.
    """
    history_path = get_call_history_path(directory)
    if not os.path.exists(history_path):
        return {}
    try:
        with open(history_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        return data if isinstance(data, dict) else {}
    except:
        return {}


def save_call_history(directory, history_dict):
    """
    Save call history atomically (write to temp then os.rename).
    """
    history_path = get_call_history_path(directory)
    temp_path = history_path + '.tmp'
    try:
        with open(temp_path, 'w', encoding='utf-8') as f:
            json.dump(history_dict, f, indent=2, ensure_ascii=False)
        os.replace(temp_path, history_path)
    except Exception as e:
        # Clean up temp file if rename failed
        try:
            os.unlink(temp_path)
        except:
            pass


def record_call_attempt(directory, phone):
    """
    Record current timestamp for a phone number.
    Strips the + prefix for normalization so all instances agree on keys.
    """
    # Normalize: strip + prefix
    key = phone.lstrip('+')
    history = load_call_history(directory)
    history[key] = time.time()
    save_call_history(directory, history)


def is_number_cooled(phone, history_dict):
    """
    Check if 70 seconds have passed since last call to this number.
    Returns True if enough time has passed (safe to call) or if never called before.
    """
    key = phone.lstrip('+')
    last_call = history_dict.get(key)
    if last_call is None:
        return True  # Never called before
    return (time.time() - last_call) >= CALL_COOLDOWN


# ============================================================================
#                     ACCOUNT INITIALIZATION (POST /init)
# ============================================================================

def init_account(acc):
    """
    Initialize a single account by calling POST /init.
    If the account already has a token, keep it.
    Otherwise, call /init to get a new session token.
    Returns the updated account dict or None on failure.
    """
    email = acc.get('email', '')
    existing_token = acc.get('x-token', '') or acc.get('token', '')
    device_id = acc.get('x-client-device-id', '') or acc.get('device_id', '')

    # If already has token and device_id, use as-is
    if existing_token and device_id:
        acc['x-token'] = existing_token
        acc['x-client-device-id'] = device_id
        return acc

    # Need to get a token via /init
    if not device_id:
        device_id = ''.join(random.choices('0123456789abcdef', k=16))
        acc['x-client-device-id'] = device_id

    # Try up to 2 times (was 3)
    for attempt in range(2):
        token = telicall_init_session(device_id)
        if token:
            acc['x-token'] = token
            acc['x-client-device-id'] = device_id
            return acc
        time.sleep(0.5)

    return None


def check_account_balance(acc, test_phone=None):
    """
    Quick balance check: Try to start a call with the account.
    If the API returns SIP credentials -> account has balance (VALID).
    If the API returns no_balance -> account is dead.
    If the API returns call_404 -> account is VALID (phone not found but has balance).
    If the API returns call_400 without balance -> account might be OK, treat as valid.
    Returns: (has_balance: bool, error_type: str or None)
    """
    call_token = acc.get('x-token', '') or acc.get('token', '')
    call_device_id = acc.get('x-client-device-id', '') or acc.get('device_id', '')

    if not call_token:
        return False, 'no_token'

    if not test_phone:
        test_phone = '+977000000000'  # Fake number - won't connect, but tests if account has balance

    result = telicall_start_call(test_phone, call_token, call_device_id)

    if result is None:
        # Timeout/network error - DON'T kill account, might be temporary
        return True, 'timeout_ok'
    elif result == 'no_balance':
        return False, 'no_balance'
    elif isinstance(result, dict) and 'error' in result:
        err = result['error']
        if err == 'call_404':
            # Phone not found - but account CAN make calls (has balance)
            return True, None
        elif err == 'call_400':
            # Bad request but not no_balance - account has balance
            return True, None
        elif err == 'call_403':
            # 403 = Forbidden for this NUMBER, NOT account dead!
            # The account has balance but the test number is blocked.
            # This is the MOST COMMON case with fake test numbers!
            return True, None
        elif err == 'call_429':
            # Rate limited - account is fine, just too many requests
            return True, None
        else:
            # Any other error - DON'T kill account on balance check!
            # The fake number might cause weird errors.
            return True, err
    elif isinstance(result, dict) and result.get('user'):
        # Got SIP credentials -> account has balance!
        return True, None
    else:
        # Got a dict but no user key - check if it has SIP-like keys
        if isinstance(result, dict) and result.get('domain'):
            return True, None
        # Unknown response - don't kill account
        return True, 'unknown'


def init_all_accounts(raw_accounts, directory):
    """
    Initialize all accounts by calling /init for those that need tokens.
    Then validates each account with a balance check (quick call test).
    Uses cache file to skip already-initialized and dead accounts.
    Does it in batches for speed.
    Returns list of ready accounts (with tokens AND balance).
    """
    # Load cache
    cache_data = load_accounts_cache(directory)
    ready, need_init, dead_from_cache = merge_cache_with_dan(cache_data, raw_accounts)

    # Check if ALL accounts are dead in cache - auto-reset
    total_from_dan = len(raw_accounts)
    if cache_data and len(dead_from_cache) >= total_from_dan and total_from_dan > 0:
        cached_dead = len(cache_data.get('dead', []))
        cached_ready = len(cache_data.get('ready', []))
        print(f"  {clr(C.BCYAN, 'Cache found!')} ({CACHE_FILENAME})")
        print(f"    {clr(C.GREEN, 'Cached ready:')} {cached_ready}")
        print(f"    {clr(C.RED, 'Cached dead:')}  {cached_dead}")
        print()
        print(f"  {clr(C.BYELLOW, 'ALL accounts dead in cache — auto-resetting cache...')}")
        # Delete cache and re-run merge automatically
        cache_path = get_cache_path(directory)
        try:
            os.unlink(cache_path)
            print(f"  {clr(C.GREEN, 'Cache deleted! Re-checking all accounts...')}")
        except:
            pass
        cache_data = None
        ready, need_init, dead_from_cache = merge_cache_with_dan(cache_data, raw_accounts)
    else:
        if SKIP_BALANCE_CHECK:
            # When skipping balance check, don't trust "dead" from cache
            # Those accounts were likely killed by a faulty balance check (call_403 on fake number)
            # Give them a second chance - they'll be tested with real numbers
            if dead_from_cache:
                print(f"  {clr(C.BYELLOW, '⚠ ' + str(len(dead_from_cache)) + ' cached dead accounts will be retried')}")
                print(f"  {clr(C.DIM, '(SKIP_BALANCE_CHECK=True - dead accounts get a second chance)')}")
                # Add dead accounts back to need_init so they get fresh tokens
                for acc in raw_accounts:
                    email = acc.get('email', '')
                    if email in dead_from_cache:
                        need_init.append(acc)
                dead_from_cache = set()  # Clear dead set
        else:
            # Add dead accounts from cache to used_emails so they're skipped
            for email in dead_from_cache:
                used_emails.add(email)

        if cache_data:
            cached_ready_count = len(cache_data.get('ready', []))
            cached_dead = len(cache_data.get('dead', []))
            cached_used = len(cache_data.get('used', []))
            print(f"  {clr(C.BCYAN, 'Cache found!')} ({CACHE_FILENAME})")
            print(f"    {clr(C.GREEN, 'Cached ready:')} {cached_ready_count}")
            print(f"    {clr(C.RED, 'Cached dead:')}  {cached_dead}")
            print(f"    {clr(C.YELLOW, 'Cached used:')}  {cached_used}")
            print()
            print(f"  {clr(C.GREEN, 'Reusing')} {len(ready)} accounts from cache {clr(C.DIM, '(skip /init)')}")
            print(f"  {clr(C.RED, 'Skipping')} {len(dead_from_cache)} dead accounts")
            print(f"  {clr(C.BLUE, 'New accounts')} needing /init: {len(need_init)}")
        else:
            print(f"  {clr(C.YELLOW, 'No cache file found')} ({CACHE_FILENAME})")
            print(f"  Accounts already with tokens: {len(ready)}")
            print(f"  Accounts needing /init: {len(need_init)}")
    print()

    if not need_init and not ready:
        # No accounts to init and no ready accounts
        save_accounts_cache(directory, ready, dead_from_cache, set())
        return ready

    # Initialize accounts in batches
    print(f"  {clr(C.BBLUE, 'Initializing accounts via POST /init ...')}")
    print()

    init_lock = threading.Lock()
    progress = {'done': 0, 'ok': 0, 'fail': 0}

    def _init_one(acc):
        result = init_account(acc)
        with init_lock:
            progress['done'] += 1
            if result:
                progress['ok'] += 1
                ready.append(result)
            else:
                progress['fail'] += 1
                failed_emails.add(acc.get('email', ''))
            # Print progress
            pct = progress['done'] * 100 // len(need_init)
            ok_s = clr(C.GREEN, f"{progress['ok']}")
            fail_s = clr(C.RED, f"{progress['fail']}")
            print(f"\r  Init: {progress['done']}/{len(need_init)} ({pct}%) | OK: {ok_s} | Failed: {fail_s}   ", end='', flush=True)

    failed_emails = set()

    # Use thread pool for parallel init
    threads = []
    for i, acc in enumerate(need_init):
        t = threading.Thread(target=_init_one, args=(acc,), daemon=True)
        t.start()
        threads.append(t)

        # Limit concurrent inits
        while sum(1 for t in threads if t.is_alive()) >= INIT_BATCH_SIZE:
            time.sleep(0.05)

    # Wait for all to finish
    for t in threads:
        t.join(timeout=15)

    print()
    print()
    ok_s = clr(C.BGREEN, f"{progress['ok']} OK")
    fail_s = clr(C.BRED, f"{progress['fail']} failed")
    print(f"  Init complete: {ok_s}, {fail_s} out of {len(need_init)}")
    print(f"  {clr(C.BCYAN, 'Total ready accounts:')} {clr(C.BGREEN, str(len(ready)))}")
    print()

    # Add failed emails to dead set
    all_dead = dead_from_cache | failed_emails

    # Save cache
    save_accounts_cache(directory, ready, all_dead, set())
    print(f"  {clr(C.GREEN, 'Cache saved to:')} {CACHE_FILENAME}")
    print()

    return ready


def validate_accounts_balance(ready_accounts, directory, test_phone=None):
    """
    Validate that accounts actually have calling balance by doing a quick
    test call attempt. This filters out accounts that have /init tokens
    but no_balance - so we don't waste time during the calling phase.
    Like the /Ff command in the bot - checks account works before using.
    test_phone: real phone number to use for balance check (from user's list)
    Returns: (valid_accounts, dead_emails_set)
    """
    if not ready_accounts:
        return [], set()

    total = len(ready_accounts)
    print(f"  {clr(C.BBLUE, 'STEP: Validating accounts (balance check via /call/outbound/start)')}")
    print(f"  {clr(C.DIM, 'Like /Ff in the bot - verify account can actually make calls')}")
    print()

    valid = []
    dead = set()
    val_lock = threading.Lock()
    progress = {'done': 0, 'ok': 0, 'no_bal': 0, 'err': 0}

    def _check_one(acc):
        email = acc.get('email', '?')
        has_bal, err = check_account_balance(acc, test_phone)
        with val_lock:
            progress['done'] += 1
            if has_bal:
                progress['ok'] += 1
                valid.append(acc)
            elif err == 'no_balance':
                progress['no_bal'] += 1
                dead.add(email)
                used_emails.add(email)
            else:
                progress['err'] += 1
                dead.add(email)
                used_emails.add(email)
            # Live progress
            pct = progress['done'] * 100 // total
            ok_s = clr(C.BGREEN, f"{progress['ok']}")
            nb_s = clr(C.BRED, f"{progress['no_bal']}")
            er_s = clr(C.RED, f"{progress['err']}")
            print(f"\r  Check: {progress['done']}/{total} ({pct}%) | {clr(C.GREEN, 'Valid:')} {ok_s} | {clr(C.RED, 'NoBal:')} {nb_s} | {clr(C.RED, 'Err:')} {er_s}   ", end='', flush=True)

    # Run balance checks in parallel (fast - just API call, no SIP)
    threads = []
    for i, acc in enumerate(ready_accounts):
        t = threading.Thread(target=_check_one, args=(acc,), daemon=True)
        t.start()
        threads.append(t)

        # Limit concurrent checks
        while sum(1 for t in threads if t.is_alive()) >= 30:
            time.sleep(0.05)

    # Wait for all to finish
    for t in threads:
        t.join(timeout=10)

    print()
    print()
    print(f"  {clr(C.BGREEN, 'Valid accounts (have balance):')} {len(valid)}")
    print(f"  {clr(C.BRED, 'No balance (dead):')}           {len(dead)}")
    print(f"  {clr(C.DIM, 'Filtered out:')}                {total - len(valid)} accounts")
    print()

    # Update cache with dead accounts from balance check
    if dead:
        save_accounts_cache(directory, valid, dead, set())
        print(f"  {clr(C.GREEN, 'Cache updated with balance results')}")
        print()

    return valid, dead


# ============================================================================
#                              SIP CLASS
# ============================================================================

class SIP:
    """SIP client for making VoIP calls via Telicall."""

    def __init__(self, u, p, d, pt, pr='tcp'):
        self.u, self.p, self.d, self.pt, self.pr = u, p, d, pt, pr
        self.lp = random.randint(50000, 60000)
        self.rtp_port = self.lp + 2
        self.tag = uuid.uuid4().hex[:8]
        self.seq = 1
        self.sk = None
        self.rs = self.rn = self.ro = self.rq = None
        self.br = self.cid = None
        self.rtp_sk = None
        self.rtp_run = False
        self.audio = []
        self.rtp_ip = None
        self.rtp_pt = None
        self.ssrc = random.randint(1000000, 9999999)
        self.rtp_seq = 0
        self.rtp_ts = 0
        self.remote_tag = None

    def conn(self):
        try:
            self.sk = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            self.sk.settimeout(SIP_CONNECT_TIMEOUT)  # WAS 30s, now 8s
            if self.pr == 'tls':
                import ssl
                context = ssl.create_default_context()
                context.check_hostname = False
                context.verify_mode = ssl.CERT_NONE
                self.sk = context.wrap_socket(self.sk, server_hostname=self.d)
            self.sk.connect((self.d, self.pt))
            return True
        except Exception:
            return False

    def _pauth(self, h):
        for k, p in [('rs', r'realm="([^"]+)"'), ('rn', r'nonce="([^"]+)"'),
                     ('ro', r'opaque="([^"]+)"'), ('rq', r'qop="([^"]+)"')]:
            m = re.search(p, h)
            if m:
                setattr(self, k, m.group(1))

    def _auth(self, method, uri):
        if not self.rs or not self.rn:
            return None
        h1 = hashlib.md5(f"{self.u}:{self.rs}:{self.p}".encode()).hexdigest()
        h2 = hashlib.md5(f"{method}:{uri}".encode()).hexdigest()
        if self.rq:
            nc, cn = "00000001", uuid.uuid4().hex[:8]
            rp = hashlib.md5(f"{h1}:{self.rn}:{nc}:{cn}:{self.rq}:{h2}".encode()).hexdigest()
            return (f'Digest username="{self.u}",realm="{self.rs}",nonce="{self.rn}",'
                    f'uri="{uri}",response="{rp}",opaque="{self.ro}",'
                    f'qop={self.rq},nc={nc},cnonce="{cn}",algorithm=MD5')
        rp = hashlib.md5(f"{h1}:{self.rn}:{h2}".encode()).hexdigest()
        return (f'Digest username="{self.u}",realm="{self.rs}",nonce="{self.rn}",'
                f'uri="{uri}",response="{rp}",opaque="{self.ro}",algorithm=MD5')

    def send(self, msg):
        try:
            if isinstance(msg, str):
                msg = msg.encode()
            self.sk.send(msg)
            return True
        except:
            return False

    def recv(self, timeout=RECV_TIMEOUT):
        try:
            self.sk.settimeout(timeout)
            data = b''
            while True:
                chunk = self.sk.recv(4096)
                if not chunk:
                    break
                data += chunk
                if b'\r\n\r\n' in data:
                    try:
                        header = data.split(b'\r\n\r\n')[0].decode('utf-8', errors='ignore')
                        cl = re.search(r'Content-Length:\s*(\d+)', header, re.IGNORECASE)
                        if cl:
                            body_start = data.find(b'\r\n\r\n') + 4
                            if len(data) >= body_start + int(cl.group(1)):
                                break
                        else:
                            break
                    except:
                        break
            return data.decode('utf-8', errors='ignore')
        except:
            return None

    def parse(self, resp):
        if not resp:
            return None
        lines = resp.split('\r\n')
        parts = lines[0].split(' ', 2)
        code = int(parts[1]) if len(parts) > 1 else 0
        headers = {}
        for line in lines[1:]:
            if ':' in line:
                k, v = line.split(':', 1)
                headers[k.strip().lower()] = v.strip()
        to_tag = None
        m = re.search(r';tag=([^;>\s]+)', headers.get('to', ''))
        if m:
            to_tag = m.group(1)
        sdp_ip, sdp_port = None, None
        if '\r\n\r\n' in resp:
            sdp = resp.split('\r\n\r\n', 1)[1]
            m = re.search(r'c=IN IP4 ([\d.]+)', sdp)
            if m:
                sdp_ip = m.group(1)
            m = re.search(r'm=audio (\d+)', sdp)
            if m:
                sdp_port = int(m.group(1))
        return {'code': code, 'headers': headers, 'to_tag': to_tag,
                'sdp_ip': sdp_ip, 'sdp_port': sdp_port}

    def register(self, auth=False):
        uri = f"sip:{self.d}"
        branch = f"z9hG4bK-{uuid.uuid4().hex[:16]}"
        call_id = f"{uuid.uuid4().hex[:16]}@{self.d}"
        msg = f"REGISTER {uri} SIP/2.0\r\n"
        msg += f"Via: SIP/2.0/{self.pr.upper()} {self.d}:{self.pt};branch={branch};rport\r\n"
        msg += f"From: <sip:{self.u}@{self.d}>;tag={self.tag}\r\n"
        msg += f"To: <sip:{self.u}@{self.d}>\r\n"
        msg += f"Call-ID: {call_id}\r\n"
        msg += f"CSeq: {self.seq} REGISTER\r\n"
        msg += f"Contact: <sip:{self.u}@{self.d}:{self.lp}>\r\n"
        msg += "Max-Forwards: 70\r\n"
        msg += "User-Agent: TelliCall/1.2.1\r\n"
        msg += "Allow: INVITE, ACK, CANCEL, BYE, OPTIONS\r\n"
        if auth and self.rn:
            a = self._auth("REGISTER", uri)
            if a:
                msg += f"Authorization: {a}\r\n"
        msg += "Content-Length: 0\r\n\r\n"
        self.seq += 1
        return self.send(msg)

    def _get_local_ip(self):
        try:
            s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
            s.connect((self.d, 80))
            ip = s.getsockname()[0]
            s.close()
            return ip
        except:
            return '0.0.0.0'

    def invite(self, number, auth=False):
        uri = f"sip:+{number}@{self.d}"
        self.br = f"z9hG4bK-{uuid.uuid4().hex[:16]}"
        self.cid = f"{uuid.uuid4().hex[:16]}@{self.d}"
        local_ip = self._get_local_ip()
        sdp = f"v=0\r\n"
        sdp += f"o=- {int(time.time())} {int(time.time())} IN IP4 {local_ip}\r\n"
        sdp += "s=TelliCall\r\n"
        sdp += f"c=IN IP4 {local_ip}\r\n"
        sdp += "t=0 0\r\n"
        sdp += f"m=audio {self.rtp_port} RTP/AVP 0 8 101\r\n"
        sdp += "a=rtpmap:0 PCMU/8000\r\n"
        sdp += "a=rtpmap:8 PCMA/8000\r\n"
        sdp += "a=rtpmap:101 telephone-event/8000\r\n"
        sdp += "a=sendrecv\r\n"
        sdp += "a=ptime:20\r\n"
        sdp_b = sdp.encode()
        msg = f"INVITE {uri} SIP/2.0\r\n"
        msg += f"Via: SIP/2.0/{self.pr.upper()} {self.d}:{self.pt};branch={self.br};rport\r\n"
        msg += f"From: <sip:{self.u}@{self.d}>;tag={self.tag}\r\n"
        msg += f"To: <sip:+{number}@{self.d}>\r\n"
        msg += f"Call-ID: {self.cid}\r\n"
        msg += f"CSeq: {self.seq} INVITE\r\n"
        msg += f"Contact: <sip:{self.u}@{self.d}:{self.lp}>\r\n"
        msg += "Max-Forwards: 70\r\n"
        msg += "User-Agent: TelliCall/1.2.1\r\n"
        msg += "Allow: INVITE, ACK, CANCEL, BYE, OPTIONS\r\n"
        msg += "Content-Type: application/sdp\r\n"
        if auth and self.rn:
            a = self._auth("INVITE", uri)
            if a:
                msg += f"Authorization: {a}\r\n"
        msg += f"Content-Length: {len(sdp_b)}\r\n\r\n"
        msg += sdp
        self.seq += 1
        return self.send(msg)

    def ack(self, number):
        msg = f"ACK sip:+{number}@{self.d} SIP/2.0\r\n"
        msg += f"Via: SIP/2.0/{self.pr.upper()} {self.d}:{self.pt};branch={self.br};rport\r\n"
        msg += f"From: <sip:{self.u}@{self.d}>;tag={self.tag}\r\n"
        msg += f"To: <sip:+{number}@{self.d}>;tag={self.remote_tag}\r\n"
        msg += f"Call-ID: {self.cid}\r\n"
        msg += f"CSeq: {self.seq} ACK\r\n"
        msg += "Max-Forwards: 70\r\n"
        msg += "Content-Length: 0\r\n\r\n"
        return self.send(msg)

    def bye(self, number):
        self.seq += 1
        branch = f"z9hG4bK-{uuid.uuid4().hex[:16]}"
        msg = f"BYE sip:+{number}@{self.d} SIP/2.0\r\n"
        msg += f"Via: SIP/2.0/{self.pr.upper()} {self.d}:{self.pt};branch={branch};rport\r\n"
        msg += f"From: <sip:{self.u}@{self.d}>;tag={self.tag}\r\n"
        msg += f"To: <sip:+{number}@{self.d}>;tag={self.remote_tag}\r\n"
        msg += f"Call-ID: {self.cid}\r\n"
        msg += f"CSeq: {self.seq} BYE\r\n"
        msg += "Max-Forwards: 70\r\n"
        msg += "Content-Length: 0\r\n\r\n"
        return self.send(msg)

    def ok(self, req):
        lines = req.split('\r\n')
        headers = {}
        for line in lines[1:]:
            if ':' in line:
                k, v = line.split(':', 1)
                headers[k.strip().lower()] = v.strip()
        msg = "SIP/2.0 200 OK\r\n"
        msg += f"Via: {headers.get('via', '')}\r\n"
        msg += f"From: {headers.get('from', '')}\r\n"
        msg += f"To: {headers.get('to', '')}\r\n"
        msg += f"Call-ID: {headers.get('call-id', '')}\r\n"
        msg += f"CSeq: {headers.get('cseq', '')}\r\n"
        msg += "Content-Length: 0\r\n\r\n"
        return self.send(msg)

    def start_rtp(self):
        try:
            self.rtp_sk = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
            self.rtp_sk.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
            try:
                self.rtp_sk.bind(('0.0.0.0', self.rtp_port))
            except OSError:
                self.rtp_sk.bind(('0.0.0.0', 0))
                self.rtp_port = self.rtp_sk.getsockname()[1]
            self.rtp_sk.settimeout(0.05)
            self.rtp_run = True
            return True
        except:
            return False

    def build_rtp(self, payload):
        pt = 0  # PCMU silence
        first = (2 << 6) | 0
        header = struct.pack('!BBHII', first, pt, self.rtp_seq, self.rtp_ts, self.ssrc)
        self.rtp_seq = (self.rtp_seq + 1) & 0xFFFF
        self.rtp_ts += 160
        return header + payload

    def send_rtp(self):
        if not self.rtp_ip or not self.rtp_pt or not self.rtp_sk:
            return False
        payload = bytes([0xFF] * 160)  # silence
        pkt = self.build_rtp(payload)
        try:
            self.rtp_sk.sendto(pkt, (self.rtp_ip, self.rtp_pt))
            return True
        except:
            return False

    def rtp_loop(self, stop_evt, dur):
        start = time.perf_counter()

        def _recv_worker():
            sock = self.rtp_sk
            try:
                sock.settimeout(1.0)
            except:
                pass
            while self.rtp_run and not stop_evt.is_set():
                try:
                    data, addr = sock.recvfrom(4096)
                    if len(data) < 12:
                        continue
                    pt = data[1] & 0x7F
                    raw = data[12:]
                    if pt == 8:
                        pcm = audioop.alaw2lin(raw, 2)
                    else:
                        pcm = audioop.ulaw2lin(raw, 2)
                    self.audio.append(pcm)
                except socket.timeout:
                    continue
                except OSError:
                    break
                except:
                    pass

        recv_thread = threading.Thread(target=_recv_worker, daemon=True)
        recv_thread.start()

        sent = 0
        next_send = start
        PTIME = 0.020

        while self.rtp_run and not stop_evt.is_set():
            now = time.perf_counter()
            if now - start >= dur:
                break
            if now >= next_send:
                self.send_rtp()
                sent += 1
                next_send = start + sent * PTIME
            remaining = next_send - time.perf_counter()
            if remaining > 0.003:
                time.sleep(remaining - 0.002)
            while time.perf_counter() < next_send:
                pass

        stop_evt.set()
        recv_thread.join(timeout=2.0)

    def stop_rtp(self):
        self.rtp_run = False
        if self.rtp_sk:
            try:
                self.rtp_sk.close()
            except:
                pass

    def close(self):
        self.stop_rtp()
        if self.sk:
            try:
                self.sk.close()
            except:
                pass


# ============================================================================
#                     AUTO ACCOUNT CREATION (from to.py, optimized)
# ============================================================================

# Egyptian IP ranges for x-real-ip header
_EG_RANGES = [
    (41, 32), (41, 33), (41, 34), (41, 35), (41, 36),
    (41, 37), (41, 38), (41, 39), (41, 40), (41, 41),
    (41, 42), (41, 43), (41, 44), (41, 45), (41, 46),
    (41, 47), (41, 48), (41, 49), (41, 50), (41, 51),
    (41, 52), (41, 53), (41, 54), (41, 55), (41, 56),
    (41, 57), (41, 58), (41, 59), (41, 60), (41, 61),
    (156, 192), (156, 193), (156, 194), (156, 195),
    (156, 196), (156, 197), (156, 198), (156, 199),
    (156, 200), (156, 201), (156, 202), (156, 203),
    (197, 32), (197, 33), (197, 34), (197, 35),
    (197, 36), (197, 37), (197, 38), (197, 39),
    (197, 40), (197, 41), (197, 42), (197, 43),
]

# IP management - UNIQUE IP for every request (like to.py)
_ip_lock = threading.Lock()
_used_ips = set()

def rand_eg_ip():
    """Generate a unique Egyptian IP every time (never reuse)."""
    with _ip_lock:
        for _ in range(50):
            a, b = random.choice(_EG_RANGES)
            c = random.randint(1, 254)
            d = random.randint(1, 254)
            ip = f"{a}.{b}.{c}.{d}"
            if ip not in _used_ips:
                _used_ips.add(ip)
                return ip
        # All used - reset and try again
        _used_ips.clear()
        a, b = random.choice(_EG_RANGES)
        c = random.randint(1, 254)
        d = random.randint(1, 254)
        return f"{a}.{b}.{c}.{d}"

def _report_ip_fail():
    pass  # Not needed with unique IP approach

def _report_ip_ok():
    pass  # Not needed with unique IP approach

# mail.tm email creation
MAIL_TM_HEADERS = {
    'Content-Type': 'application/json',
    'Accept': 'application/json',
    'Origin': 'https://mail.tm',
    'Referer': 'https://mail.tm/en/',
    'User-Agent': 'Mozilla/5.0 (Linux; Android 11; Infinix X698) AppleWebKit/537.36',
}
MAIL_TM_DOMAIN = "wshu.net"
_mail_tm_lock = threading.Lock()  # Needed to avoid duplicate emails

def create_mail_tm():
    """Create a temporary email on mail.tm (wshu.net) - with lock to prevent duplicates"""
    for attempt in range(5):  # Increased from 3 like to.py
        with _mail_tm_lock:
            name = ''.join(random.choices(string.ascii_lowercase + string.digits, k=12))
            email_addr = f"{name}@{MAIL_TM_DOMAIN}"
            password = 'TmpP@ss' + ''.join(random.choices(string.digits, k=4))
            try:
                r = requests.post('https://api.mail.tm/accounts',
                    json={'address': email_addr, 'password': password},
                    headers=MAIL_TM_HEADERS, timeout=15)
                if r.status_code in [200, 201]:
                    r2 = requests.post('https://api.mail.tm/token',
                        json={'address': email_addr, 'password': password},
                        headers=MAIL_TM_HEADERS, timeout=10)
                    if r2.status_code == 200:
                        jwt = r2.json()['token']
                        return {'email': email_addr, 'token': jwt, 'password': password}
                elif r.status_code == 429:
                    time.sleep(10 * (attempt + 1))  # Like to.py: 10s, 20s, 30s...
                else:
                    time.sleep(2)  # Like to.py
            except Exception:
                time.sleep(2)  # Like to.py
    return None

def get_otp_from_mail(jwt):
    """Wait for OTP from mail.tm inbox - polling every 2s, max 60s (like to.py)"""
    deadline = time.time() + 60  # Like to.py: 60 seconds
    while time.time() < deadline:
        try:
            r = requests.get('https://api.mail.tm/messages',
                headers={'Authorization': f'Bearer {jwt}', **MAIL_TM_HEADERS},
                timeout=8)  # reduced from 10
            if r.status_code == 200:
                data = r.json()
                messages = data.get('hydra:member', []) if isinstance(data, dict) else data
                for msg in messages:
                    msg_id = msg.get('id', '')
                    if msg_id:
                        r2 = requests.get(f'https://api.mail.tm/messages/{msg_id}',
                            headers={'Authorization': f'Bearer {jwt}', **MAIL_TM_HEADERS},
                            timeout=8)
                        if r2.status_code == 200:
                            full_msg = r2.json()
                            content = full_msg.get('text', '') or str(full_msg)
                            if 'teli' in content.lower():
                                m = re.search(r'\b(\d{6})\b', content)
                                if m:
                                    return m.group(1)
        except:
            pass
        time.sleep(2)  # Like to.py: 2s polling
    return None

def init_session_with_ip():
    """Init session with unique Egyptian IP - EXACTLY like to.py"""
    ip = rand_eg_ip()
    device = ''.join(random.choices('0123456789abcdef', k=16))
    h = {
        "host": "api.telicall.com",
        "x-request-id": str(uuid.uuid4()),
        "user-agent": "Dalvik/2.1.0",
        "x-app-version": "1.2.1",
        "x-client-device-id": device,
        "x-lang": "en", "x-os": "android", "x-os-version": "11",
        "x-req-timestamp": str(int(time.time() * 1000)),
        "x-req-signature": "-1",
        "content-type": "application/json",
        "x-token": "",
        "x-currency": "EGP",
        "x-real-ip": ip,
    }
    body = {
        "countryCode": "eg", "deviceName": "Infinix X698",
        "notificationToken": "", "oldToken": "",
        "peerKey": str(random.randint(100, 999)),
        "timeZone": "Africa/Cairo", "localizationKey": ""
    }
    try:
        # Refresh request-specific headers (like to.py)
        h["x-request-id"] = str(uuid.uuid4())
        h["x-req-timestamp"] = str(int(time.time() * 1000))
        r = requests.post(f"{API_URL}/init", json=body, headers=h, timeout=10)
        if r.status_code == 200:
            tok = r.json().get('result', {}).get('token')
            if tok:
                h["x-token"] = tok
                return tok, device, h  # Return headers too (like to.py)
            else:
                print(f"    {clr(C.RED, '⚠ init [' + ip + ']: no token in response')}", flush=True)
        else:
            print(f"    {clr(C.RED, '⚠ init [' + ip + ']: HTTP ' + str(r.status_code))}", flush=True)
    except Exception as e:
        print(f"    {clr(C.RED, '⚠ init [' + ip + ']: ' + str(e)[:50])}", flush=True)
    return None, None, None

def send_verify_email(email, headers_or_token, device_id=None):
    """Send verification email - supports both old (token,device) and new (headers) style"""
    # Support both calling conventions:
    # Old: send_verify_email(email, token, device_id)
    # New: send_verify_email(email, headers_dict)  <- like to.py
    if isinstance(headers_or_token, dict):
        # New style: headers dict passed directly (like to.py)
        h = headers_or_token.copy()
        h["x-request-id"] = str(uuid.uuid4())
        h["x-req-timestamp"] = str(int(time.time() * 1000))
    else:
        # Old style: token + device_id
        h = get_headers(token=headers_or_token, device_id=device_id)
    try:
        r = requests.post(f"{API_URL}/auth/send-email", json={'email': email}, headers=h, timeout=10)
        if r.status_code == 200:
            return r.json().get('result', {}).get('reference')
        else:
            try:
                err = r.json().get('meta', {}).get('errorMessage', r.text[:80])
            except:
                err = r.text[:80]
            print(f"    {clr(C.RED, '⚠ send_verify: ' + str(r.status_code) + ' | ' + str(err)[:50])}", flush=True)
    except Exception as e:
        print(f"    {clr(C.RED, '⚠ send_verify: ' + str(e)[:50])}", flush=True)
    return None

def verify_otp_code(ref, code, headers_or_token, device_id=None):
    """Verify OTP code - supports both old (token,device) and new (headers) style"""
    if isinstance(headers_or_token, dict):
        h = headers_or_token.copy()
        h["x-request-id"] = str(uuid.uuid4())
        h["x-req-timestamp"] = str(int(time.time() * 1000))
    else:
        h = get_headers(token=headers_or_token, device_id=device_id)
    try:
        r = requests.post(f"{API_URL}/auth/verify-identity",
                          json={'reference': ref, 'code': str(code)},
                          headers=h, timeout=10)
        if r.status_code == 200:
            return r.json().get('result', {}).get('user')  # Like to.py - returns user data
        else:
            try:
                err = r.json().get('meta', {}).get('errorMessage', r.text[:80])
            except:
                err = r.text[:80]
            print(f"    {clr(C.RED, '⚠ verify_otp: ' + str(r.status_code) + ' | ' + str(err)[:50])}", flush=True)
    except Exception as e:
        print(f"    {clr(C.RED, '⚠ verify_otp: ' + str(e)[:50])}", flush=True)
    return None

def create_one_account_fast():
    """Create a single new Telicall account - EXACTLY like to.py logic.
    Key difference: passes full headers dict (with x-real-ip, x-currency) 
    to all API calls, not just token+device_id."""
    # 1. Create email
    mail = create_mail_tm()
    if not mail:
        return None

    # 2. Init session with unique Egyptian IP (returns headers dict like to.py)
    tok, device, headers = init_session_with_ip()
    if not tok:
        return None

    # 3. Send verification - pass full headers dict (like to.py)
    ref = send_verify_email(mail['email'], headers)
    if not ref:
        return None

    # 4. Wait for OTP
    otp = get_otp_from_mail(mail['token'])
    if not otp:
        return None

    # 5. Verify OTP - pass full headers dict (like to.py)
    user = verify_otp_code(ref, otp, headers)
    if not user:
        return None

    # Success!
    return {
        'email': mail['email'],
        'x-token': tok,
        'x-client-device-id': device,
        'password': mail.get('password', ''),
    }

def auto_create_accounts(directory, count=5, max_retries=3):
    """
    Create new accounts when accounts run out.
    Creates `count` accounts in FULL PARALLEL (no lock, no stagger).
    Returns list of new VALID account dicts.
    Also saves them to Dan.json and cache.
    """
    if not AUTO_CREATE_ACCOUNTS:
        return []

    all_new = []
    
    for batch in range(max_retries):
        print(f"\n  {clr(C.BYELLOW, '⚡ Auto-creating ' + str(count) + ' accounts (batch ' + str(batch+1) + '/' + str(max_retries) + ')...')}")
        print(f"  {clr(C.DIM, 'Each thread uses a different IP via x-real-ip header')}")

        new_accounts = []
        create_lock = threading.Lock()
        progress = {'done': 0, 'ok': 0, 'fail': 0}

        def _create_one(idx):
            # create_one_account_fast() calls init_session_with_ip() which
            # calls rand_eg_ip() - unique IP for every account creation
            acc = create_one_account_fast()
            with create_lock:
                progress['done'] += 1
                if acc:
                    progress['ok'] += 1
                    new_accounts.append(acc)
                    print(f"    {clr(C.GREEN, '✓ #' + str(progress['ok']) + '/' + str(progress['done']) + ':')} {acc['email'][:30]}", flush=True)
                else:
                    progress['fail'] += 1
                    print(f"    {clr(C.RED, '✗ #' + str(progress['done']) + ' failed')}", flush=True)

        # Create ALL in parallel immediately (no stagger!)
        threads = []
        for i in range(count):
            t = threading.Thread(target=_create_one, args=(i,), daemon=True)
            t.start()
            threads.append(t)

        # Wait for all threads
        for t in threads:
            t.join(timeout=90)  # max 90s per account creation

        print(f"  {clr(C.CYAN, 'Batch result:')} {clr(C.GREEN, str(progress['ok']) + ' OK')} / {clr(C.RED, str(progress['fail']) + ' failed')}")

        if new_accounts:
            all_new.extend(new_accounts)

            # Save new accounts to Dan.json (create it if it doesn't exist)
            dan_path = None
            for candidate in ["Dan.json", "dan.json"]:
                full_path = os.path.join(directory, candidate)
                if os.path.exists(full_path):
                    dan_path = full_path
                    break

            if not dan_path:
                # Create new Dan.json file
                dan_path = os.path.join(directory, "Dan.json")

            try:
                # Load existing accounts (or start fresh if file doesn't exist)
                raw_accounts = []
                if os.path.exists(dan_path) and os.path.getsize(dan_path) > 0:
                    try:
                        raw_accounts = load_dan_json(dan_path)
                    except Exception:
                        raw_accounts = []  # Corrupted file - start fresh

                for acc in new_accounts:
                    raw_accounts.append({
                        'email': acc['email'],
                        'x-client-device-id': acc.get('x-client-device-id', ''),
                        'x-token': acc.get('x-token', ''),
                        'created': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    })
                key = hashlib.sha256(ACCOUNTS_PASSWORD.encode()).digest()
                json_str = json.dumps(raw_accounts, indent=2, ensure_ascii=False)
                raw = json_str.encode('utf-8')
                enc = base64.b64encode(bytes([raw[i] ^ key[i % len(key)] for i in range(len(raw))]))
                with open(dan_path, 'wb') as f:
                    f.write(enc)
                print(f"  {clr(C.GREEN, 'Saved ' + str(len(new_accounts)) + ' accounts to Dan.json')}")
            except Exception as e:
                print(f"  {clr(C.YELLOW, 'Warning: Could not save to Dan.json:')} {e}")

            # Add to global accounts list
            for acc in new_accounts:
                accounts.append(acc)

            # Skip balance check when SKIP_BALANCE_CHECK is True (much faster!)
            # Accounts will be tested with real numbers - bad ones will die naturally
            if SKIP_BALANCE_CHECK:
                valid_new = new_accounts
                with stats_lock:
                    stats["auto_created"] += len(valid_new)
                print(f"  {clr(C.BGREEN, '✓ ' + str(len(valid_new)) + ' accounts created!')} (skip balance check = faster)")
                print(f"  {clr(C.CYAN, 'Accounts ready:')} {len(valid_new)} | {clr(C.CYAN, 'Total accounts:')} {len(accounts)}")
            else:
                # Quick balance validation (skip detailed progress - just check)
                valid_new = []
                dead_new_emails = set()
                for acc in new_accounts:
                    has_bal, _ = check_account_balance(acc)
                    if has_bal:
                        valid_new.append(acc)
                    else:
                        dead_new_emails.add(acc.get('email', ''))
                        used_emails.add(acc.get('email', ''))

                # Remove dead from accounts list
                if dead_new_emails:
                    accounts[:] = [a for a in accounts if a.get('email', '') not in dead_new_emails]
                    all_new = [a for a in all_new if a.get('email', '') not in dead_new_emails]

                with stats_lock:
                    stats["auto_created"] += len(valid_new)

                print(f"  {clr(C.BGREEN, '✓ ' + str(len(valid_new)) + ' valid accounts created!')} ({len(dead_new_emails)} had no balance)")
                print(f"  {clr(C.CYAN, 'Accounts ready:')} {len(valid_new)} | {clr(C.CYAN, 'Total accounts:')} {len(accounts)}")

            # Update cache
            cache_dir = getattr(mark_used, 'cache_dir', directory)
            save_accounts_cache(cache_dir, accounts, used_emails, set())
            
            if valid_new:
                return valid_new
            # All created accounts had no balance - try again
            print(f"  {clr(C.YELLOW, 'All new accounts had no balance - retrying...')}")
        
        # Failed this batch, wait before retry
        if batch < max_retries - 1:
            wait = min(3 * (batch + 1), 15)  # 3s, 6s, 9s... max 15s
            print(f"  {clr(C.YELLOW, 'Batch failed, retrying in ' + str(wait) + 's...')}")
            time.sleep(wait)
    
    print(f"  {clr(C.BRED, 'Failed to create accounts after ' + str(max_retries) + ' attempts!')}")
    return []


# ============================================================================
#                          CALL EXECUTION
# ============================================================================

def do_single_call(phone, dur, info, status_cb=None, just_detect=None):
    """
    Execute a single SIP call and return the result.
    Returns: (result_str, from_number)
      result_str: 'answered_ok', 'no_answer', 'declined', 'failed', 'not_found', 'answered_short'
      from_number: the caller ID used
      status_cb: optional callback(status_str) for live updates
      just_detect: override JUST_DETECT_ANSWER for this call (None = use global)
    """
    call_start = time.time()

    def _status(msg):
        if status_cb:
            status_cb(msg)

    def _timed_out():
        return (time.time() - call_start) > MAX_CALL_TIME

    sip = SIP(info['user'], info['pass'], info['domain'], info['port'], info['proto'])
    sip._from_num = str(info.get('from', '')).replace('+', '')

    _status("connecting")
    if not sip.conn():
        return ('failed', '')

    if _timed_out():
        sip.close()
        return ('failed', sip._from_num or '')

    # REGISTER
    _status("registering")
    sip.register(auth=False)
    r = sip.recv(RECV_TIMEOUT)  # WAS 10s, now 4s
    if r:
        p = sip.parse(r)
        if p and p['code'] == 401:
            sip._pauth(p['headers'].get('www-authenticate', ''))
            sip.register(auth=True)
            sip.recv(RECV_TIMEOUT)  # WAS 10s, now 4s

    if _timed_out():
        sip.close()
        return ('failed', sip._from_num or '')

    # INVITE
    _status("inviting")
    num = phone.replace('+', '')
    sip.invite(num, auth=False)
    r = sip.recv(RECV_TIMEOUT)  # WAS 10s, now 4s

    if not r:
        sip.close()
        return ('failed', sip._from_num or '')

    p = sip.parse(r)
    if not p or p['code'] != 401:
        # Got something unexpected (not 401 challenge)
        code = p['code'] if p else 0
        if code == 200:
            # Direct 200 OK without auth - call answered immediately
            sip.remote_tag = p['to_tag']
            sdp_ip = p['sdp_ip']
            sdp_port = p['sdp_port']
            if sdp_ip and sdp_port:
                sip.ack(num)
                # Short call check then return answered
                sip.close()
                return ('answered_ok', sip._from_num or '')
        sip.close()
        return ('failed', sip._from_num or '')

    if _timed_out():
        sip.close()
        return ('failed', sip._from_num or '')

    # INVITE with auth
    sip._pauth(p['headers'].get('www-authenticate', ''))
    sip.seq -= 1
    sip.invite(num, auth=True)

    _status("ringing")
    ringing_started = False
    call_answered = False
    sdp_ip = sdp_port = None

    for i in range(RINGING_TIMEOUT):  # WAS 140 (70s), now 40 (20s)
        if _timed_out():
            sip.close()
            return ('failed', sip._from_num or '')

        r = sip.recv(0.5)
        if r:
            p = sip.parse(r)
            code = p['code'] if p else 0

            if code == 100:
                pass
            elif code == 180 or code == 183:
                ringing_started = True
                _status("ringing")
            elif code == 200:
                call_answered = True
                sip.remote_tag = p['to_tag']
                sdp_ip = p['sdp_ip']
                sdp_port = p['sdp_port']

                if not sdp_ip or not sdp_port:
                    sip.close()
                    return ('failed', sip._from_num or '')

                sip.ack(num)

                _just_detect = just_detect if just_detect is not None else JUST_DETECT_ANSWER
                if _just_detect:
                    # Just detect answer - hang up immediately, don't stay on the call
                    time.sleep(0.2)
                    sip.bye(num)
                    sip.close()
                    return ('answered_ok', sip._from_num or '')

                # Check for instant BYE (quick - only 1s)
                sip.sk.settimeout(0.2)
                instant_bye = False
                for _check in range(INSTANT_BYE_CHECKS):  # WAS 10 (2s), now 3 (0.6s)
                    try:
                        chk = sip.sk.recv(4096)
                        if chk:
                            chk_str = chk.decode('utf-8', errors='ignore')
                            if 'BYE ' in chk_str:
                                instant_bye = True
                                sip.ok(chk_str)
                                break
                    except:
                        pass

                if instant_bye:
                    sip.close()
                    return ('declined', sip._from_num or '')

                break

            elif code == 486:
                sip.close()
                return ('declined', sip._from_num or '')
            elif code == 487:
                sip.close()
                return ('declined', sip._from_num or '')
            elif code == 603:
                sip.close()
                return ('declined', sip._from_num or '')
            elif code == 404:
                sip.close()
                return ('not_found', sip._from_num or '')
            elif code == 408:
                sip.close()
                return ('no_answer', sip._from_num or '')
            elif code == 403:
                sip.close()
                return ('failed', sip._from_num or '')
            elif code == 480:
                sip.close()
                return ('no_answer', sip._from_num or '')
            elif code == 488:
                sip.close()
                return ('failed', sip._from_num or '')
            elif code >= 400:
                sip.close()
                return ('no_answer', sip._from_num or '')

    if not call_answered:
        sip.close()
        r = 'no_answer' if ringing_started else 'failed'
        return (r, sip._from_num or '')

    # Call was answered - stay in the call for CALL_DURATION seconds
    _status("answered")
    sip.rtp_ip = sdp_ip if sdp_ip else sip.d
    sip.rtp_pt = sdp_port if sdp_port else 5004

    # Track that this call is now in the "talking" phase
    with in_call_lock:
        global in_call_count
        in_call_count += 1

    stop_evt = threading.Event()
    if sip.start_rtp():
        rt = threading.Thread(target=sip.rtp_loop, args=(stop_evt, dur), daemon=True)
        rt.start()

    time.sleep(0.3)
    start_time = time.time()
    deadline = start_time + dur
    call_ended = False

    sip.sk.settimeout(0.1)
    while time.time() < deadline:
        try:
            chk = sip.sk.recv(4096)
            if chk:
                chk_str = chk.decode('utf-8', errors='ignore')
                if 'BYE ' in chk_str:
                    first = chk_str.strip().split('\r\n')[0] if '\r\n' in chk_str else chk_str.strip().split('\n')[0]
                    if first.startswith('BYE ') or '\r\nBYE ' in chk_str or '\nBYE ' in chk_str:
                        sip.ok(chk_str)
                        call_ended = True
                        break
        except:
            pass

    actual = min(int(time.time() - start_time), dur)
    stop_evt.set()
    sip.stop_rtp()

    # Decrement in-call counter
    with in_call_lock:
        in_call_count -= 1

    if not call_ended:
        sip.bye(num)

    sip.close()

    status = 'answered_ok' if actual >= MIN_ANSWERED_DURATION else 'answered_short'
    return (status, sip._from_num or '')


# ============================================================================
#                          FILE READERS
# ============================================================================

def find_xlsx_files(directory):
    """Find all .xlsx files in the given directory."""
    xlsx_files = []
    for f in os.listdir(directory):
        if f.endswith('.xlsx') and not f.startswith('~$'):
            xlsx_files.append(os.path.join(directory, f))
    return sorted(xlsx_files)


def read_phones_from_xlsx(filepath):
    """Read phone numbers from an .xlsx file."""
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb.active
    phones = []
    number_col = None

    for row in ws.iter_rows(min_row=1, max_row=3, values_only=False):
        for cell in row:
            if cell.value and str(cell.value).strip().lower() in ('number', 'phone', 'mobile', 'numbers', 'phones'):
                number_col = cell.column
                break
        if number_col:
            break

    for row in ws.iter_rows(min_row=2, values_only=False):
        if number_col:
            cell = row[number_col - 1] if len(row) >= number_col else None
        else:
            cell = row[0] if row else None

        if cell and cell.value:
            val = str(cell.value).strip()
            cleaned = re.sub(r'[^\d+]', '', val)
            if cleaned and len(cleaned) >= 7:
                if not cleaned.startswith('+'):
                    cleaned = '+' + cleaned
                phones.append(cleaned)

    wb.close()
    return phones


def load_dan_json(filepath):
    """Load and decrypt Dan.json account file."""
    key = hashlib.sha256(ACCOUNTS_PASSWORD.encode()).digest()
    raw_b64 = open(filepath, 'rb').read()
    raw = base64.b64decode(raw_b64)
    text = bytes([raw[i] ^ key[i % len(key)] for i in range(len(raw))]).decode('utf-8')
    return json.loads(text)


# ============================================================================
#                        DISPLAY HELPERS
# ============================================================================

def print_banner():
    print()
    print(clr(C.BCYAN, "  " + "=" * 60))
    print(clr(C.BWHITE, "     ____                _      ____              _    "))
    print(clr(C.BWHITE, "    |  _ \\ __ _ _ __ __| | ___/ ___|  __ _ _ __ | |__ "))
    print(clr(C.BWHITE, "    | |_) / _` | '__/ _` |/ _ \\___ \\ / _` | '_ \\| '_ \\"))
    print(clr(C.BWHITE, "    |  _ < (_| | | | (_| |  __/___) | (_| | |_) | | | |"))
    print(clr(C.BWHITE, "    |_| \\_\\__,_|_|  \\__,_|\\___|____/ \\__,_| .__/|_| |_|"))
    print(clr(C.BWHITE, "                                          |_|          "))
    print(clr(C.BCYAN, "  " + "=" * 60))
    print(clr(C.DIM, "     Bulk SIP Calling Tool | Powered by Telicall API"))
    print(clr(C.BCYAN, "  " + "=" * 60))
    print()


def add_call_log(msg, color=None):
    """Add a message to the call log for display."""
    ts = datetime.now().strftime('%H:%M:%S')
    with call_log_lock:
        if color and COLOR:
            call_log.append(f"{C.DIM}[{ts}]{C.RST} {color}{msg}{C.RST}")
        else:
            call_log.append(f"{clr(C.DIM, '[' + ts + ']')} {msg}")
        if len(call_log) > MAX_LOG_LINES:
            call_log.pop(0)


def print_stats_line():
    """Print a single-line live status update."""
    with stats_lock:
        a = stats["answered"]
        na = stats["no_answer"]
        b = stats["busy"]
        f = stats["failed"]
        nf = stats["not_found"]
        af = stats["api_fail"]
        au = stats["accounts_used"]
        total = stats["total_numbers"]
    with done_lock:
        d = len(done_numbers)
    with active_lock:
        ac = active_calls
    with in_call_lock:
        ic = in_call_count
    with number_lock:
        qlen = len(number_queue)
    rem = count_remaining()

    effective = d + ic
    pct = (effective * 100 // total) if total > 0 else 0

    # Color-coded stats
    d_s = clr(C.BGREEN, f"{d}") if d > 0 else f"{d}"
    ic_s = clr(C.CYAN, f"{ic}") if ic > 0 else f"{ic}"
    a_s = clr(C.BGREEN, f"{a}") if a > 0 else f"{a}"
    na_s = clr(C.YELLOW, f"{na}") if na > 0 else f"{na}"
    b_s = clr(C.BYELLOW, f"{b}") if b > 0 else f"{b}"
    f_s = clr(C.BRED, f"{f}") if f > 0 else f"{f}"
    af_s = clr(C.RED, f"{af}") if af > 0 else f"{af}"

    print(
        f"\r  {d_s}+{ic_s}/{total} ({pct}%) | "
        f"{clr(C.BLUE, 'Act')}:{ac} | "
        f"{clr(C.CYAN, 'Talk')}:{ic_s} | "
        f"{clr(C.GREEN, 'Ans')}:{a_s} | "
        f"{clr(C.YELLOW, 'NoA')}:{na_s} | "
        f"{clr(C.BYELLOW, 'Bsy')}:{b_s} | "
        f"{clr(C.RED, 'Fail')}:{f_s} | "
        f"{clr(C.RED, 'API')}:{af_s} | "
        f"Left:{rem}   ",
        end='', flush=True
    )


def print_live_display():
    """Print a multi-line live display with stats and recent call log."""
    with stats_lock:
        a = stats["answered"]
        na = stats["no_answer"]
        b = stats["busy"]
        f = stats["failed"]
        nf = stats["not_found"]
        af = stats["api_fail"]
        au = stats["accounts_used"]
        total = stats["total_numbers"]
    with done_lock:
        d = len(done_numbers)
    with active_lock:
        ac = active_calls
    with in_call_lock:
        ic = in_call_count
    with number_lock:
        qlen = len(number_queue)
    rem = count_remaining()

    # Effective progress = done + in_call (they're answered, just talking)
    effective = d + ic
    pct = (effective * 100 // total) if total > 0 else 0
    pct_bar_len = 30
    filled = pct * pct_bar_len // 100
    bar = clr(C.BGREEN, "=" * filled) + clr(C.DIM, "-" * (pct_bar_len - filled))

    # Build output
    lines = []
    lines.append(f"  {clr(C.BCYAN, chr(9552) * 55)}")
    lines.append(f"  {clr(C.BWHITE, 'PROGRESS')} [{bar}] {clr(C.BGREEN, str(d))}{clr(C.DIM, '+')}{clr(C.CYAN, str(ic))}/{total} ({pct}%)")

    # Stats with colors
    ans_s = clr(C.BGREEN, f"{a:>3}") if a > 0 else clr(C.DIM, "  0")
    ic_s = clr(C.CYAN, f"{ic:>3}") if ic > 0 else clr(C.DIM, "  0")
    na_s = clr(C.YELLOW, f"{na:>3}") if na > 0 else clr(C.DIM, "  0")
    bu_s = clr(C.BYELLOW, f"{b:>3}") if b > 0 else clr(C.DIM, "  0")
    fa_s = clr(C.BRED, f"{f:>3}") if f > 0 else clr(C.DIM, "  0")
    nf_s = clr(C.MAGENTA, f"{nf:>3}") if nf > 0 else clr(C.DIM, "  0")
    af_s = clr(C.RED, f"{af:>3}") if af > 0 else clr(C.DIM, "  0")

    lines.append(f"  {clr(C.GREEN, 'Ans')}{ans_s}  {clr(C.CYAN, 'Talk')}{ic_s}  {clr(C.YELLOW, 'NoA')}{na_s}  {clr(C.BYELLOW, 'Bsy')}{bu_s}  {clr(C.RED, 'Fail')}{fa_s}  {clr(C.RED, 'API')}{af_s}")
    lines.append(f"  {clr(C.BLUE, 'Active:')} {clr(C.BWHITE, f'{ac:>3}')}  {clr(C.CYAN, 'Queue:')} {qlen:>3}  {clr(C.DIM, 'Accounts:')} {rem}")

    # Recent call log
    with call_log_lock:
        log_copy = list(call_log[-8:])

    if log_copy:
        lines.append(f"  {clr(C.DIM, chr(9472) * 55)}")
        for entry in log_copy:
            lines.append(f"  {entry}")

    lines.append(f"  {clr(C.BCYAN, chr(9552) * 55)}")

    # Calculate how many lines to clear
    num_lines = len(lines)

    # Move cursor up to overwrite
    if hasattr(print_live_display, 'prev_lines'):
        prev = print_live_display.prev_lines
        if prev > 0:
            sys.stdout.write(f"\033[{prev}A")
    print_live_display.prev_lines = num_lines

    # Clear and write
    for i, line in enumerate(lines):
        sys.stdout.write(f"\r\033[K{line}\n")
    sys.stdout.flush()


def print_final_results(phones, answered_list, start_time):
    """Print final results and save to file."""
    elapsed = time.time() - start_time
    remaining = count_remaining()

    print()
    print()
    print(f"  {clr(C.BCYAN, '=' * 55)}")
    print(f"  {clr(C.BWHITE, 'FINAL RESULTS')}")
    print(f"  {clr(C.BCYAN, '=' * 55)}")
    print()
    with stats_lock:
        s = stats  # shorthand
        print(f"  Total Numbers:        {s['total_numbers']}")
        print(f"  Actually Called:      {clr(C.BWHITE, str(s['total_numbers'] - s['skipped']))}")
        print(f"  {clr(C.RED, 'Skipped (no accounts):')}  {clr(C.BRED, str(s['skipped']))}")
        print()
        print(f"  {clr(C.BGREEN, 'Answered:')}           {clr(C.BGREEN, str(s['answered']))}")
        print(f"  {clr(C.YELLOW, 'No Answer:')}          {s['no_answer']}")
        print(f"  {clr(C.BYELLOW, 'Busy:')}               {s['busy']}")
        print(f"  {clr(C.RED, 'Failed:')}            {s['failed']}")
        print(f"  {clr(C.MAGENTA, 'Not Found:')}         {s['not_found']}")
        print(f"  {clr(C.YELLOW, 'Answered & Hung Up:')} {s['answered_short']}")
        print(f"  {clr(C.RED, 'API Failures:')}       {s['api_fail']}")
        print()
        print(f"  {clr(C.DIM, 'Accounts Used:')}        {s['accounts_used']}")
        if s['auto_created'] > 0:
            print(f"  {clr(C.CYAN, 'Accounts Auto-Created:')} {clr(C.BCYAN, str(s['auto_created']))}")
    print(f"  {clr(C.DIM, 'Accounts Remaining:')}   {remaining}")
    print(f"  Time Elapsed:         {elapsed:.1f}s ({elapsed/60:.1f}min)")

    # Success rate
    called = s['total_numbers'] - s['skipped']
    if called > 0:
        rate = s['answered'] * 100 // called
        print(f"  Success Rate:         {clr(C.BGREEN, str(rate) + '%')} ({s['answered']}/{called} called numbers)")
    print()

    # Save answered numbers to file
    if answered_list:
        unique = list(dict.fromkeys(answered_list))
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filepath = f"answered_numbers_{timestamp}.txt"
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(f"# Fox Caller - Answered Numbers\n")
            f.write(f"# Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"# Total: {len(unique)} numbers\n\n")
            for ph in unique:
                f.write(f"{ph}\n")
        print(f"  {clr(C.GREEN, 'Answered numbers saved to:')} {filepath}")
        print()
        print(f"  {clr(C.BGREEN, 'Answered Numbers')} ({len(unique)}):")
        print()
        for i, ph in enumerate(unique, 1):
            print(f"    {clr(C.GREEN, str(i) + '.')} {ph}")
    else:
        print(f"  {clr(C.YELLOW, 'No numbers were answered.')}")


# ============================================================================
#                    SINGLE MODE - Parallel (4 accounts per number)
# ============================================================================

def process_number_parallel(phone, num_parallel=None):
    """
    Process one phone number with multiple accounts SIMULTANEOUSLY.
    Fires num_parallel SIP calls at once with different accounts.
    The first account that gets answered wins; the rest are released.
    Returns True if answered, False otherwise.
    """
    if num_parallel is None:
        num_parallel = PARALLEL_ACCOUNTS

    # Get up to num_parallel accounts
    accs = get_next_accounts(num_parallel)

    if not accs:
        # Try auto-creating accounts
        if AUTO_CREATE_ACCOUNTS:
            directory = getattr(mark_used, 'cache_dir', '.')
            new = auto_create_accounts(directory, ACCOUNT_CREATE_BATCH)
            if new:
                global account_index
                account_index = 0
                accs = get_next_accounts(num_parallel)

        if not accs:
            print(f"    {clr(C.BRED, 'No accounts available!')}")
            return False

    # Record call attempt for cooldown tracking
    directory = getattr(mark_used, 'cache_dir', '.')
    record_call_attempt(directory, phone)

    # Fire parallel calls
    call_results = {}
    results_lock = threading.Lock()
    answered_event = threading.Event()  # Set when any call is answered

    def _make_call(idx, acc):
        email = acc.get('email', '???')
        email_short = email[:25]

        # If another thread already got an answer, skip this call
        if answered_event.is_set():
            with results_lock:
                call_results[idx] = ('skipped', '', email, '')
            return

        info, email_used = start_call_with_account(phone, acc)

        if info is None:
            # Network/timeout - account might still be good
            with results_lock:
                call_results[idx] = ('api_timeout', '', email_used, 'timeout')
            return
        elif info == 'no_balance':
            # Account truly dead
            with results_lock:
                call_results[idx] = ('no_balance', '', email_used, 'no_balance')
            return
        elif isinstance(info, dict) and "error" in info:
            err = info.get("error", "")
            if err == 'call_403':
                # 403 = Number blocked, NOT account dead - mark as blocked number
                with results_lock:
                    call_results[idx] = ('blocked', '', email_used, 'call_403')
                return
            elif err == 'call_404':
                # Number not found
                with results_lock:
                    call_results[idx] = ('not_found_api', '', email_used, 'call_404')
                return
            else:
                with results_lock:
                    call_results[idx] = ('api_fail', '', email_used, err)
                return

        # If another thread already got an answer, don't make SIP call
        if answered_event.is_set():
            with results_lock:
                call_results[idx] = ('skipped', '', email, '')
            return

        result, from_num = do_single_call(phone, CALL_DURATION, info)

        if result in ('answered_ok', 'answered_short'):
            answered_event.set()  # Signal other threads to stop

        with results_lock:
            call_results[idx] = (result, from_num, email_used, '')

    threads = []
    for i, acc in enumerate(accs):
        email_short = acc.get('email', '???')[:20]
        print(f"    {clr(C.DIM, '[' + str(i+1) + '/' + str(len(accs)) + ']')} {email_short}...", flush=True)
        t = threading.Thread(target=_make_call, args=(i, acc), daemon=True)
        t.start()
        threads.append(t)
        time.sleep(0.15)  # slight stagger to avoid API rate limits

    # Wait for all threads to complete
    for t in threads:
        t.join(timeout=35)

    # Process results
    answered = False
    all_blocked = True  # Track if ALL calls were blocked (403)
    for idx in sorted(call_results.keys()):
        result, from_num, email_used, err = call_results[idx]

        if result in ('answered_ok', 'answered_short'):
            answered = True
            all_blocked = False
            mark_used(email_used, status='used')
            print(f"    {clr(C.BGREEN, chr(10004) + ' ANSWERED')} via {email_used[:25]} (from: +{from_num})")
            with stats_lock:
                stats["answered"] += 1
        elif result == 'blocked':
            # call_403 - number is blocked, NOT account dead
            all_blocked = all_blocked  # stays True
            print(f"    {clr(C.RED, chr(10007) + ' Blocked (403)')} number can't be called")
        elif result == 'not_found_api':
            all_blocked = False
            with stats_lock:
                stats["not_found"] += 1
            print(f"    {clr(C.MAGENTA, chr(10007) + ' Not Found (404)')}")
        elif result == 'no_balance':
            all_blocked = False
            mark_used(email_used, status='dead')
            print(f"    {clr(C.RED, chr(10007) + ' No Balance')} account dead")
        elif result == 'api_timeout':
            all_blocked = False
            with stats_lock:
                stats["api_fail"] += 1
            print(f"    {clr(C.RED, chr(10007) + ' API Timeout')}")
        elif result == 'api_fail':
            all_blocked = False
            with stats_lock:
                stats["api_fail"] += 1
            print(f"    {clr(C.RED, chr(10007) + ' API:')} {err}")
        elif result == 'no_answer':
            all_blocked = False
            with stats_lock:
                stats["no_answer"] += 1
            print(f"    {clr(C.YELLOW, chr(10007) + ' No Answer')}")
        elif result == 'declined':
            all_blocked = False
            with stats_lock:
                stats["busy"] += 1
            print(f"    {clr(C.BYELLOW, chr(10007) + ' Busy')}")
        elif result == 'not_found':
            all_blocked = False
            with stats_lock:
                stats["not_found"] += 1
            print(f"    {clr(C.MAGENTA, chr(10007) + ' Not Found')}")
        elif result == 'failed':
            all_blocked = False
            with stats_lock:
                stats["failed"] += 1
            print(f"    {clr(C.RED, chr(10007) + ' SIP Failed')}")
        elif result == 'skipped':
            pass  # Another call already answered
        else:
            all_blocked = False

    # If ALL calls were blocked (403), mark this number as done
    # so we don't waste more accounts on it
    if all_blocked and not answered:
        with stats_lock:
            stats["failed"] += 1
        print(f"    {clr(C.BRED, 'Number BLOCKED')} - all accounts got 403 for this number")

    return answered


def run_single_mode(phones):
    """Single mode: Random number selection, 4 accounts per number simultaneously."""
    global account_index
    total = len(phones)
    answered_list = []
    start_time = time.time()

    with stats_lock:
        stats["total_numbers"] = total

    # Shuffle numbers for RANDOM selection (not sequential!)
    # This prevents multiple instances from calling the same number
    shuffled_phones = list(phones)
    random.shuffle(shuffled_phones)

    print()
    print(f"  {clr(C.BBLUE, 'Starting Parallel Mode...')}")
    print(f"  Numbers: {total} | Accounts: {count_remaining()}")
    print(f"  {clr(C.DIM, str(PARALLEL_ACCOUNTS) + ' accounts per number | Detect answer only (no waiting)')}")
    print(f"  {clr(C.DIM, 'Random order | Cooldown: ' + str(CALL_COOLDOWN) + 's between same number')}")
    print()

    auto_fail_streak = 0
    processed = 0
    skipped_cooldown = 0
    i = 0

    while i < len(shuffled_phones):
        phone = shuffled_phones[i]
        processed += 1

        # Check cooldown - skip if called recently by another instance
        directory = getattr(mark_used, 'cache_dir', '.')
        history = load_call_history(directory)
        if not is_number_cooled(phone, history):
            # Number was called recently - skip it
            skipped_cooldown += 1
            with stats_lock:
                stats["skipped"] += 1
            i += 1
            continue

        # Check if we have accounts available
        remaining_accounts = count_remaining()
        if remaining_accounts == 0:
            # Try auto-creating accounts
            if AUTO_CREATE_ACCOUNTS:
                new = auto_create_accounts(directory, ACCOUNT_CREATE_BATCH)
                if new:
                    auto_fail_streak = 0
                    account_index = 0
                    remaining_accounts = count_remaining()
                    if remaining_accounts == 0:
                        auto_fail_streak += 1
                        if auto_fail_streak >= 10:
                            print(f"\n  {clr(C.BRED, 'Auto-create failed 10 times!')} Stopping.")
                            for p in shuffled_phones[i:]:
                                with stats_lock:
                                    stats["skipped"] += 1
                            break
                        wait_time = min(10 * auto_fail_streak, 60)
                        print(f"\n  {clr(C.YELLOW, 'Auto-create failed. Retrying in ' + str(wait_time) + 's...')}")
                        time.sleep(wait_time)
                        continue
                else:
                    auto_fail_streak += 1
                    if auto_fail_streak >= 10:
                        print(f"\n  {clr(C.BRED, 'Auto-create failed 10 times!')} Stopping.")
                        for p in shuffled_phones[i:]:
                            with stats_lock:
                                stats["skipped"] += 1
                        break
                    wait_time = min(10 * auto_fail_streak, 60)
                    print(f"\n  {clr(C.YELLOW, 'Auto-create failed (' + str(auto_fail_streak) + '/10). Retrying in ' + str(wait_time) + 's...')}")
                    time.sleep(wait_time)
                    continue
            else:
                print(f"\n  {clr(C.BRED, 'All accounts consumed!')} Stopping at number {processed}/{total}.")
                for p in shuffled_phones[i:]:
                    with stats_lock:
                        stats["skipped"] += 1
                break

        num_acc = min(PARALLEL_ACCOUNTS, count_remaining())
        print(f"  {clr(C.BWHITE, '-- ' + str(processed) + '/' + str(total) + ':')} {phone}  {clr(C.DIM, '(' + str(num_acc) + ' accounts)')}")

        answered = process_number_parallel(phone)
        if answered:
            answered_list.append(phone)
        with stats_lock:
            stats["total_processed"] += 1
        i += 1

    print_final_results(phones, answered_list, start_time)
    return answered_list


# ============================================================================
#                    AUTO MODE - Parallel Per-Account Streaming
# ============================================================================

def pick_next_number():
    """
    Pick the next available number from the shared queue using RANDOM selection.
    Skips numbers that are already done (answered or max retries).
    Checks call history cooldown - skips numbers called within last 70 seconds.
    Returns phone number string or None if queue is empty / all on cooldown.
    """
    # Load call history for cooldown check
    directory = getattr(mark_used, 'cache_dir', '.')
    history = load_call_history(directory)

    with number_lock:
        if not number_queue:
            return None

        # Convert deque to list for random access
        phone_list = list(number_queue)
        number_queue.clear()

        cooled_numbers = []   # numbers safe to call
        on_cooldown = []      # numbers still on cooldown

        for phone in phone_list:
            with done_lock:
                if phone in done_numbers:
                    continue  # skip done numbers
            if is_number_cooled(phone, history):
                cooled_numbers.append(phone)
            else:
                on_cooldown.append(phone)

        # Put on-cooldown numbers back in the queue
        for phone in on_cooldown:
            number_queue.append(phone)

        if not cooled_numbers:
            return None  # All remaining numbers are on cooldown

        # Randomly pick from cooled numbers
        phone = random.choice(cooled_numbers)

        # Put the rest of the cooled numbers back in the queue
        for p in cooled_numbers:
            if p != phone:
                number_queue.append(p)

        return phone


def requeue_number(phone, result_str=None):
    """
    Put a number back in the queue for retry (no_answer, busy).
    Only requeue if max retries not exceeded.
    Returns True if requeued, False if max retries exceeded (number is done).
    result_str: the call result that caused this requeue (for final classification)
    """
    # Track last result for this number
    if result_str:
        with number_last_result_lock:
            number_last_result[phone] = result_str

    with number_attempts_lock:
        number_attempts[phone] = number_attempts.get(phone, 0) + 1
        if number_attempts[phone] >= MAX_RETRIES_PER_NUMBER:
            # Max retries exceeded - classify based on last result
            with number_last_result_lock:
                last = number_last_result.get(phone, 'failed')
            with stats_lock:
                if last == 'no_answer':
                    stats["no_answer"] += 1
                elif last in ('declined', 'busy'):
                    stats["busy"] += 1
                elif last == 'not_found':
                    stats["not_found"] += 1
                else:
                    stats["failed"] += 1
                stats["total_processed"] += 1
            with done_lock:
                done_numbers.add(phone)
            add_call_log(f"{clr(C.RED, 'MAX_RETRY')} {phone} (last: {last})")
            return False
    with number_lock:
        number_queue.append(phone)
    return True


def mark_number_done(phone):
    """Mark a number as fully processed (answered or given up)."""
    with done_lock:
        done_numbers.add(phone)
    with stats_lock:
        stats["total_processed"] += 1


def init_number_queue(phones):
    """Initialize the shared number queue with shuffled phone numbers."""
    shuffled = list(phones)
    random.shuffle(shuffled)  # Random order, not sequential
    with number_lock:
        number_queue.clear()
        for p in shuffled:
            number_queue.append(p)
    with number_attempts_lock:
        number_attempts.clear()
    with number_last_result_lock:
        number_last_result.clear()
    with done_lock:
        done_numbers.clear()


def auto_account_worker(worker_id, phones, stop_event):
    """
    Worker thread for auto mode.
    Keeps getting NEW accounts from the pool.
    When one account dies, it gets another and continues.
    When a number is answered, the account is consumed, gets a new one.
    Only stops when: no more accounts OR no more numbers.
    """
    global active_calls, account_index

    while not stop_event.is_set():
        # Get the next available account from the pool
        acc = get_next_account()
        if acc is None:
            # No more accounts available - this worker is done
            break

        email = acc.get('email', '???')
        email_short = email[:20]

        # Inner loop: use this account for calls until it dies or is consumed
        while not stop_event.is_set():
            # Pick the next number from the queue
            phone = pick_next_number()
            if phone is None:
                # No more numbers to call - this worker is done
                return

            with active_lock:
                active_calls += 1

            # Live log: starting call
            add_call_log(f"{clr(C.BLUE, 'DIAL')} {phone} <- {email_short}")

            # Record call attempt for cooldown tracking
            directory = getattr(mark_used, 'cache_dir', '.')
            record_call_attempt(directory, phone)

            # Start the call via Telicall API
            info, email_used = start_call_with_account(phone, acc)

            if info is None or info == 'no_balance' or (isinstance(info, dict) and "error" in info):
                # API failed to start call
                err_detail = ""
                if isinstance(info, dict):
                    err_detail = info.get('error', '')
                elif info == 'no_balance':
                    err_detail = 'no_balance'
                elif info is None:
                    err_detail = 'timeout'

                # Decide: is this an account problem or a number problem?
                if err_detail == 'call_403':
                    # 403 = region/number block - NOT account dead!
                    # Mark this number as failed (don't retry), keep account alive
                    mark_number_done(phone)
                    with number_last_result_lock:
                        number_last_result[phone] = 'failed'
                    with active_lock:
                        active_calls -= 1
                    with stats_lock:
                        stats["failed"] += 1
                    add_call_log(f"{clr(C.RED, 'BLOCKED')} {phone} (403 region block) <- {email_short}")
                    # Keep going with same account - pick next number
                    continue

                elif err_detail == 'call_404':
                    # Number not found - account is fine, skip this number
                    mark_number_done(phone)
                    with number_last_result_lock:
                        number_last_result[phone] = 'not_found'
                    with active_lock:
                        active_calls -= 1
                    with stats_lock:
                        stats["not_found"] += 1
                    add_call_log(f"{clr(C.MAGENTA, 'NOT_FOUND')} {phone} <- {email_short}")
                    # Keep going with same account
                    continue

                elif err_detail == 'call_400':
                    # Bad request - might be number format issue, skip number
                    requeue_number(phone, result_str='api_fail')
                    with active_lock:
                        active_calls -= 1
                    with stats_lock:
                        stats["api_fail"] += 1
                    add_call_log(f"{clr(C.RED, 'API_ERR')} {phone} (400) <- {email_short}")
                    # Don't kill account for 400 - keep going
                    continue

                elif err_detail == 'no_balance':
                    # Account truly dead - no balance
                    requeue_number(phone, result_str='api_fail')
                    with active_lock:
                        active_calls -= 1
                    with stats_lock:
                        stats["api_fail"] += 1
                    add_call_log(f"{clr(C.RED, 'NO_BAL')} {phone} <- {email_short}")
                    # Account is dead - break inner loop, get new account
                    break

                else:
                    # Timeout or other error - might be temporary
                    requeue_number(phone, result_str='api_fail')
                    with active_lock:
                        active_calls -= 1
                    with stats_lock:
                        stats["api_fail"] += 1
                    add_call_log(f"{clr(C.RED, 'API_ERR')} {phone} ({err_detail}) <- {email_short}")
                    # For timeouts, account might still be good - keep going
                    continue

            # Execute the SIP call with live status callback
            phone_short = phone[-7:] if len(phone) > 7 else phone

            def make_status_cb(ph, em):
                last_status = {'s': ''}
                def cb(status):
                    # Only log important status changes
                    if status != last_status['s']:
                        last_status['s'] = status
                        if status == 'ringing':
                            add_call_log(f"{clr(C.CYAN, 'RING')} {ph} <- {em}")
                        elif status == 'answered':
                            add_call_log(f"{clr(C.BGREEN, 'UP!')} {ph} <- {em}")
                return cb

            result, from_num = do_single_call(phone, CALL_DURATION, info, status_cb=make_status_cb(phone_short, email_short))

            # Process result
            if result == 'answered_ok':
                # SUCCESS - mark account used, number is done
                mark_used(email_used, status='used')
                mark_number_done(phone)
                with stats_lock:
                    stats["answered"] += 1
                with answered_lock:
                    answered_phones.append(phone)
                with active_lock:
                    active_calls -= 1
                add_call_log(f"{clr(C.BGREEN, 'ANSWERED')} {phone} via {email_short} ({CALL_DURATION}s)")
                # Account is consumed after a successful call - break inner loop, get new account
                break

            elif result == 'answered_short':
                # Answered but hung up quickly - count as answered, account consumed
                mark_used(email_used, status='used')
                mark_number_done(phone)
                with stats_lock:
                    stats["answered"] += 1
                    stats["answered_short"] += 1
                with answered_lock:
                    answered_phones.append(phone)
                with active_lock:
                    active_calls -= 1
                add_call_log(f"{clr(C.GREEN, 'ANS_SHORT')} {phone} via {email_short}")
                break

            elif result == 'declined':
                # Busy - put number back for retry, this account tries next number
                with stats_lock:
                    stats["busy"] += 1
                requeue_number(phone, result_str='declined')
                with active_lock:
                    active_calls -= 1
                add_call_log(f"{clr(C.BYELLOW, 'BUSY')} {phone} <- {email_short}")
                # Keep going with same account - pick next number

            elif result == 'no_answer':
                # No answer - put number back for retry by another account
                with stats_lock:
                    stats["no_answer"] += 1
                requeue_number(phone, result_str='no_answer')
                with active_lock:
                    active_calls -= 1
                add_call_log(f"{clr(C.YELLOW, 'NO_ANS')} {phone} <- {email_short}")
                # Keep going with same account - pick next number

            elif result == 'not_found':
                # Number not found - don't retry, mark as done
                mark_number_done(phone)
                with stats_lock:
                    stats["not_found"] += 1
                with active_lock:
                    active_calls -= 1
                add_call_log(f"{clr(C.MAGENTA, 'NOT_FOUND')} {phone} <- {email_short}")
                # Keep going with same account - pick next number

            elif result == 'failed':
                # SIP failure (403, 488) - usually number issue, NOT account dead
                # Account can still call other numbers
                requeue_number(phone, result_str='failed')
                with stats_lock:
                    stats["failed"] += 1
                with active_lock:
                    active_calls -= 1
                add_call_log(f"{clr(C.RED, 'FAILED')} {phone} <- {email_short}")
                # Keep going with same account - pick next number
                continue

            else:
                # Unknown result - account might be dead
                mark_used(email_used, status='dead')
                requeue_number(phone, result_str='failed')
                with stats_lock:
                    stats["failed"] += 1
                with active_lock:
                    active_calls -= 1
                add_call_log(f"{clr(C.RED, 'UNKNOWN')} {phone} <- {email_short}")
                break

    return


def run_sequential_mode(phones, num_workers=3):
    """
    Concurrent Sequential mode: N workers call numbers simultaneously.
    Each answered call stays alive for 62 seconds (after the person answers),
    then hangs up and immediately calls the next available number.
    Loops through all numbers infinitely in rounds.

    Key differences from auto mode:
    - N concurrent calls (configurable via --workers, default 3)
    - Call stays alive for 62 seconds AFTER ANSWER then we hang up
    - After call ends, worker immediately calls next available number
    - Uses fox_call_history.json to avoid re-calling same number within 70s
    - No Dan.json dependency - auto-creates accounts as needed
    """
    global account_index, accounts

    total = len(phones)
    start_time = time.time()
    directory = getattr(mark_used, 'cache_dir', '.')

    with stats_lock:
        stats["total_numbers"] = total

    print()
    print(f"  {clr(C.BBLUE, 'Starting Concurrent Sequential Mode...')}")
    print(f"  Numbers: {total} | Workers: {num_workers} | Call duration: {SEQ_CALL_DURATION}s after answer")
    print(f"  Cooldown: {CALL_COOLDOWN}s | No delay between calls")
    print(f"  {clr(C.DIM, f'{num_workers}x RING -> answer -> stay 62s -> hang up -> RING next -> repeat')}")
    print()

    # Shared number queue
    phone_queue = deque(list(phones))
    random.shuffle(phone_queue)
    queue_lock = threading.Lock()
    refill_lock = threading.Lock()

    # Shared 403 tracking
    consecutive_403 = [0]
    c403_lock = threading.Lock()
    MAX_CONSECUTIVE_403 = 30

    # Stop flag
    stop_flag = [False]

    # Round tracking
    round_num = [1]

    def get_next_phone():
        """Get next available phone from queue (not on cooldown)."""
        with queue_lock:
            history = load_call_history(directory)
            tried = 0
            total_in_q = len(phone_queue)
            while tried < total_in_q:
                phone = phone_queue.popleft()
                if is_number_cooled(phone, history):
                    return phone
                else:
                    phone_queue.append(phone)
                    tried += 1
            return None  # All on cooldown or queue empty

    def refill_queue():
        """Refill queue with all numbers (shuffled) for new round."""
        with refill_lock:
            with queue_lock:
                phone_queue.clear()
                shuffled = list(phones)
                random.shuffle(shuffled)
                phone_queue.extend(shuffled)
            round_num[0] += 1
            print(f"  {clr(C.BCYAN, f'--- Round {round_num[0]} ---')}")

    def worker(worker_id):
        """Worker thread - makes calls one at a time, stays 62s per answered call."""
        while not stop_flag[0]:
            # Get next phone number
            phone = get_next_phone()
            if phone is None:
                # All on cooldown or queue empty
                # Check if ALL numbers are on cooldown
                history = load_call_history(directory)
                all_cooled = all(not is_number_cooled(p, history) for p in phones)
                if all_cooled:
                    # Find minimum wait time
                    min_wait = CALL_COOLDOWN
                    for p in phones:
                        key = p.lstrip('+')
                        last = history.get(key, 0)
                        wait = CALL_COOLDOWN - (time.time() - last)
                        if wait < min_wait:
                            min_wait = wait
                    wait_time = max(int(min_wait) + 1, 2)
                    print(f"  {clr(C.CYAN, f'[W{worker_id}] All numbers on cooldown - waiting {wait_time}s...')}")
                    time.sleep(wait_time)
                # Refill queue for next round
                refill_queue()
                continue

            # Get an account
            acc = get_next_account()
            if acc is None:
                if AUTO_CREATE_ACCOUNTS:
                    print(f"  {clr(C.BYELLOW, f'[W{worker_id}] No accounts - auto-creating...')}")
                    new = auto_create_accounts(directory, ACCOUNT_CREATE_BATCH, max_retries=3)
                    if new:
                        with account_lock:
                            accounts.extend(new)
                            account_index = len(accounts) - len(new)
                        acc = get_next_account()
                    else:
                        print(f"  {clr(C.BRED, f'[W{worker_id}] Failed to create accounts!')}")
                        stop_flag[0] = True
                        return
                if acc is None:
                    print(f"  {clr(C.BRED, f'[W{worker_id}] No accounts available!')}")
                    stop_flag[0] = True
                    return

            email = acc.get('email', '???')
            email_short = email[:20]
            phone_short = phone[-7:] if len(phone) > 7 else phone

            # Record call attempt
            record_call_attempt(directory, phone)

            # Start the call via Telicall API
            info, email_used = start_call_with_account(phone, acc)

            if info is None or info == 'no_balance' or (isinstance(info, dict) and "error" in info):
                err_detail = ""
                if isinstance(info, dict):
                    err_detail = info.get('error', '')
                elif info == 'no_balance':
                    err_detail = 'no_balance'
                elif info is None:
                    err_detail = 'timeout'

                if err_detail == 'call_403':
                    with c403_lock:
                        consecutive_403[0] += 1
                        if consecutive_403[0] >= MAX_CONSECUTIVE_403:
                            print(f"  {clr(C.BRED, f'{MAX_CONSECUTIVE_403} consecutive 403s - region blocked! Stopping.')}")
                            stop_flag[0] = True
                            return
                    with stats_lock:
                        stats["failed"] += 1
                    print(f"  {clr(C.RED, f'[W{worker_id}] BLOCKED')} {phone} (403) <- {email_short}")
                    continue

                elif err_detail == 'call_404':
                    with c403_lock:
                        consecutive_403[0] = 0
                    with stats_lock:
                        stats["not_found"] += 1
                    print(f"  {clr(C.MAGENTA, f'[W{worker_id}] NOT_FOUND')} {phone} <- {email_short}")
                    continue

                elif err_detail == 'no_balance':
                    mark_used(email_used, status='dead')
                    with c403_lock:
                        consecutive_403[0] = 0
                    with stats_lock:
                        stats["api_fail"] += 1
                    print(f"  {clr(C.RED, f'[W{worker_id}] NO_BAL')} {phone} <- {email_short}")
                    continue

                else:
                    with c403_lock:
                        consecutive_403[0] = 0
                    with stats_lock:
                        stats["api_fail"] += 1
                    print(f"  {clr(C.RED, f'[W{worker_id}] API_ERR')} {phone} ({err_detail}) <- {email_short}")
                    continue

            # Reset 403 counter on successful API call
            with c403_lock:
                consecutive_403[0] = 0

            # Execute the SIP call - KEEP IT ALIVE for 62s after answer!
            print(f"  {clr(C.CYAN, f'[W{worker_id}] RING')} {phone} <- {email_short}")

            call_start = time.time()
            result, from_num = do_single_call(
                phone,
                SEQ_CALL_DURATION,       # Stay in call 62 seconds AFTER answer
                info,
                just_detect=False         # DON'T hang up after answer - stay 62s!
            )
            call_duration = time.time() - call_start

            # Process result
            if result in ('answered_ok', 'answered_short'):
                mark_used(email_used, status='used')
                with stats_lock:
                    stats["answered"] += 1
                with answered_lock:
                    answered_phones.append(phone)
                dur_str = f"{int(call_duration)}s"
                print(f"  {clr(C.BGREEN, f'[W{worker_id}] ANSWERED')} {phone} via {email_short} ({dur_str})")

            elif result == 'declined':
                with stats_lock:
                    stats["busy"] += 1
                print(f"  {clr(C.BYELLOW, f'[W{worker_id}] BUSY')} {phone} <- {email_short}")

            elif result == 'no_answer':
                with stats_lock:
                    stats["no_answer"] += 1
                print(f"  {clr(C.YELLOW, f'[W{worker_id}] NO_ANS')} {phone} <- {email_short}")

            elif result == 'not_found':
                with stats_lock:
                    stats["not_found"] += 1
                print(f"  {clr(C.MAGENTA, f'[W{worker_id}] NOT_FOUND')} {phone} <- {email_short}")

            elif result == 'failed':
                with stats_lock:
                    stats["failed"] += 1
                print(f"  {clr(C.RED, f'[W{worker_id}] FAILED')} {phone} <- {email_short}")

            else:
                with stats_lock:
                    stats["failed"] += 1
                print(f"  {clr(C.RED, f'[W{worker_id}] UNKNOWN')} {phone} <- {email_short}")

            # Print running stats
            with stats_lock:
                ans = stats["answered"]
                na = stats["no_answer"]
                bs = stats["busy"]
                fl = stats["failed"]
                nf = stats["not_found"]
            print(f"  {clr(C.DIM, f'[W{worker_id}] Stats:')} {clr(C.GREEN, str(ans) + ' Ans')} | {clr(C.YELLOW, str(na) + ' NoA')} | {clr(C.BYELLOW, str(bs) + ' Bsy')} | {clr(C.RED, str(fl) + ' Fail')} | {clr(C.MAGENTA, str(nf) + ' NF')}")
            print()

            # No delay - immediately call next number
            # (answered calls already lasted 62s, no need to wait)

    # Launch worker threads
    print(f"  {clr(C.BCYAN, f'Launching {num_workers} workers...')}")
    print()
    threads = []
    for i in range(num_workers):
        t = threading.Thread(target=worker, args=(i + 1,), daemon=True)
        t.start()
        threads.append(t)
        time.sleep(0.3)  # Stagger workers slightly to avoid API burst

    # Monitor loop - print stats periodically
    last_stats_time = time.time()
    while not stop_flag[0]:
        if all(not t.is_alive() for t in threads):
            break
        # Print stats every 30 seconds
        if time.time() - last_stats_time >= 30:
            _print_seq_stats(start_time, total)
            last_stats_time = time.time()
        time.sleep(1)

    # Wait for workers to finish
    for t in threads:
        t.join(timeout=5)

    # Final stats
    _print_seq_stats(start_time, total)


def _print_seq_stats(start_time, total):
    """Print sequential mode statistics."""
    elapsed = time.time() - start_time
    with stats_lock:
        ans = stats["answered"]
        na = stats["no_answer"]
        bs = stats["busy"]
        fl = stats["failed"]
        nf = stats["not_found"]
        api_f = stats["api_fail"]

    total_calls = ans + na + bs + fl + nf + api_f
    mins = int(elapsed // 60)
    secs = int(elapsed % 60)
    print()
    print(f"  {clr(C.BBLUE, '═══ Stats ═══')}")
    print(f"  Time: {mins}m {secs}s | Total calls: {total_calls}")
    print(f"  {clr(C.BGREEN, 'Answered:')} {ans} | {clr(C.YELLOW, 'No Answer:')} {na} | {clr(C.BYELLOW, 'Busy:')} {bs}")
    print(f"  {clr(C.RED, 'Failed:')} {fl} | {clr(C.MAGENTA, 'Not Found:')} {nf} | {clr(C.RED, 'API Fail:')} {api_f}")
    if total_calls > 0:
        rate = ans * 100 / total_calls
        print(f"  {clr(C.BCYAN, f'Answer Rate:')} {rate:.1f}%")
    print()


def run_auto_mode(phones):
    """
    Auto mode: Fixed pool of workers processes all numbers.
    Workers keep getting new accounts when their current one dies.
    Workers keep picking new numbers from the shared queue.
    No batch rounds - continuous streaming.
    """
    global account_index
    total = len(phones)
    start_time = time.time()

    with stats_lock:
        stats["total_numbers"] = total

    # Initialize the shared number queue
    init_number_queue(phones)

    # Reset account index
    account_index = 0

    print()
    print(f"  {clr(C.BBLUE, 'Starting Auto Mode (Parallel Streaming)...')}")
    print(f"  Numbers: {total} | Ready accounts: {len(accounts)}")
    print(f"  Workers: {NUM_WORKERS} | Max retries/number: {MAX_RETRIES_PER_NUMBER}")
    print(f"  {clr(C.DIM, 'SIP timeout: ' + str(MAX_CALL_TIME) + 's | Ringing: ' + str(RINGING_TIMEOUT * 0.5) + 's')}")
    print()

    # Create a stop event
    stop_event = threading.Event()

    # Launch persistent worker threads
    workers = []
    actual_workers = min(NUM_WORKERS, len(accounts), total)

    print(f"  {clr(C.BCYAN, 'Launching')} {actual_workers} worker threads...")
    print()

    for i in range(actual_workers):
        t = threading.Thread(
            target=auto_account_worker,
            args=(i, phones, stop_event),
            daemon=True
        )
        t.start()
        workers.append(t)

    # Give workers a moment to start
    time.sleep(0.5)

    # Monitor progress while workers are running
    # This loop keeps going until ALL numbers are done or truly unreachable
    last_display = 0
    auto_create_fail_streak = 0       # consecutive auto-create failures
    MAX_AUTO_CREATE_RETRIES = 10      # max consecutive failures before giving up
    cooldown_wait_start = None        # track when we started waiting for cooldowns

    while True:
        now = time.time()
        with number_lock:
            qlen = len(number_queue)
        with active_lock:
            ac = active_calls
        with done_lock:
            dcount = len(done_numbers)
        rem = count_remaining()

        # Update display every 1 second
        if now - last_display >= 1.0:
            print_live_display()
            last_display = now

        # Check if all work is done
        all_workers_done = all(not t.is_alive() for t in workers)
        if dcount >= total:
            break
        if qlen == 0 and ac == 0 and all_workers_done:
            break

        # Check if we have numbers but no active calls
        if qlen > 0 and ac == 0:
            # Count how many numbers are NOT on cooldown (actually callable)
            directory = getattr(mark_used, 'cache_dir', '.')
            history = load_call_history(directory)
            callable_count = 0
            with number_lock:
                for p in number_queue:
                    if p not in done_numbers and is_number_cooled(p, history):
                        callable_count += 1

            if callable_count == 0 and qlen > 0:
                # All numbers are on cooldown - wait for cooldowns to expire
                if cooldown_wait_start is None:
                    cooldown_wait_start = now
                    add_call_log(f"{clr(C.CYAN, 'COOLDOWN')} {qlen} numbers waiting ({CALL_COOLDOWN}s)")
                # Wait for cooldowns, don't give up
                if now - cooldown_wait_start > CALL_COOLDOWN + 10:
                    # Cooldown should have expired by now - force expire old entries
                    cooldown_wait_start = None
                time.sleep(2)
                # Try to relaunch workers if accounts available
                if rem > 0 and all_workers_done:
                    new_workers = min(NUM_WORKERS, rem, qlen)
                    for i in range(new_workers):
                        t = threading.Thread(
                            target=auto_account_worker,
                            args=(len(workers) + i, phones, stop_event),
                            daemon=True
                        )
                        t.start()
                        workers.append(t)
                    time.sleep(1)
                continue
            else:
                cooldown_wait_start = None

            if rem == 0:
                # No accounts available - try auto-creating accounts
                if AUTO_CREATE_ACCOUNTS:
                    # Calculate how many accounts we need
                    needed = min(qlen, NUM_WORKERS)  # at most NUM_WORKERS at a time
                    create_count = max(ACCOUNT_CREATE_BATCH, needed)
                    directory = getattr(mark_used, 'cache_dir', '.')
                    new = auto_create_accounts(directory, create_count)
                    if new:
                        auto_create_fail_streak = 0
                        # Reset account index so workers can pick up new accounts
                        account_index = 0
                        # Relaunch workers with new accounts
                        new_workers = min(NUM_WORKERS, len(new), callable_count if callable_count > 0 else qlen)
                        if new_workers > 0:
                            for i in range(new_workers):
                                t = threading.Thread(
                                    target=auto_account_worker,
                                    args=(len(workers) + i, phones, stop_event),
                                    daemon=True
                                )
                                t.start()
                                workers.append(t)
                            time.sleep(1)
                        continue
                    else:
                        auto_create_fail_streak += 1
                        if auto_create_fail_streak >= MAX_AUTO_CREATE_RETRIES:
                            # Too many failures - give up
                            print(f"\n  {clr(C.BRED, 'Auto-create failed ' + str(MAX_AUTO_CREATE_RETRIES) + ' times in a row!')}")
                            with number_lock:
                                leftover = list(number_queue)
                                number_queue.clear()
                            skipped_count = 0
                            for p in leftover:
                                if p not in done_numbers:
                                    mark_number_done(p)
                                    skipped_count += 1
                            with stats_lock:
                                stats["skipped"] += skipped_count
                            print(f"  {clr(C.BRED, str(skipped_count) + ' numbers were never called (skipped).')}")
                            break
                        else:
                            # Wait and retry auto-creation
                            wait_time = min(10 * auto_create_fail_streak, 60)
                            print(f"\n  {clr(C.YELLOW, 'Auto-create failed (' + str(auto_create_fail_streak) + '/' + str(MAX_AUTO_CREATE_RETRIES) + '). Retrying in ' + str(wait_time) + 's...')}")
                            time.sleep(wait_time)
                            continue

                # No auto-create - mark remaining as skipped
                with number_lock:
                    leftover = list(number_queue)
                    number_queue.clear()
                skipped_count = 0
                for p in leftover:
                    if p not in done_numbers:
                        mark_number_done(p)
                        skipped_count += 1
                with stats_lock:
                    stats["skipped"] += skipped_count
                print(f"\n\n  {clr(C.BRED, 'No more accounts!')} {skipped_count} numbers were never called (skipped).")
                break
            elif all_workers_done:
                # Workers died but accounts remain - relaunch workers
                alive_workers = sum(1 for t in workers if t.is_alive())
                if alive_workers == 0 and rem > 0:
                    new_workers = min(NUM_WORKERS, rem, qlen)
                    print(f"\n  {clr(C.BYELLOW, 'Restarting')} {new_workers} workers ({rem} accounts left, {qlen} numbers in queue)")
                    for i in range(new_workers):
                        t = threading.Thread(
                            target=auto_account_worker,
                            args=(len(workers) + i, phones, stop_event),
                            daemon=True
                        )
                        t.start()
                        workers.append(t)
                    time.sleep(1)
                    continue

        time.sleep(0.3)

    # Wait for all workers to finish
    for t in workers:
        t.join(timeout=3)

    # Mark any remaining unprocessed numbers with their last result
    with number_lock:
        remaining_in_queue = list(number_queue)
    for p in remaining_in_queue:
        if p not in done_numbers:
            # Check last result for proper classification
            with number_last_result_lock:
                last = number_last_result.get(p, None)
            if last == 'no_answer':
                with stats_lock:
                    stats["no_answer"] += 1
            elif last in ('declined', 'busy'):
                with stats_lock:
                    stats["busy"] += 1
            elif last == 'not_found':
                with stats_lock:
                    stats["not_found"] += 1
            elif last and last != 'api_fail':
                with stats_lock:
                    stats["failed"] += 1
            else:
                with stats_lock:
                    stats["skipped"] += 1
            mark_number_done(p)

    # Count any remaining unprocessed numbers
    with done_lock:
        done_count = len(done_numbers)
    if done_count < total:
        unprocessed = total - done_count
        print(f"\n  {clr(C.BRED, str(unprocessed))} numbers could not be reached (accounts exhausted)")

    print_final_results(phones, answered_phones, start_time)
    return answered_phones


# ============================================================================
#                          MAIN ENTRY POINT
# ============================================================================

def main():
    """Usage:
    python3 fox_caller.py <file.xlsx> [--seq] [--workers N] [directory]

    Modes:
      (default)  Auto mode: parallel workers, fast calling
      --seq      Concurrent sequential mode: N workers, stay 62s per answered call
      --workers  Number of concurrent workers in seq mode (default: 3)
    """
    print_banner()

    # --- Step 1: Parse arguments ---
    directory = os.getcwd()
    selected_file = None
    seq_mode = False
    num_workers = 3

    argv = sys.argv[1:]
    i = 0
    while i < len(argv):
        arg = argv[i]
        if arg == '--seq':
            seq_mode = True
        elif arg == '--workers':
            if i + 1 < len(argv) and argv[i + 1].isdigit():
                num_workers = int(argv[i + 1])
                num_workers = max(1, min(num_workers, 20))  # Clamp 1-20
                i += 1
        elif arg.endswith(('.xlsx', '.txt', '.csv')):
            selected_file = os.path.abspath(arg)
        elif not arg.startswith('-'):
            directory = os.path.abspath(arg)
        i += 1

    # If no file specified, find first .xlsx in directory
    if not selected_file:
        xlsx_files = find_xlsx_files(directory)
        if xlsx_files:
            selected_file = xlsx_files[0]
            print(f"  {clr(C.GREEN, 'Auto-selected:')} {os.path.basename(selected_file)}")
        else:
            print(f"  {clr(C.BRED, 'No .xlsx files found!')}")
            print(f"  Usage: python3 fox_caller.py <file.xlsx> [--seq] [directory]")
            sys.exit(1)
    else:
        if not os.path.exists(selected_file):
            print(f"  {clr(C.BRED, 'File not found:')} {selected_file}")
            sys.exit(1)
        directory = os.path.dirname(selected_file)

    mode_str = clr(C.BCYAN, f'SEQUENTIAL ({num_workers} workers)') if seq_mode else clr(C.BGREEN, 'AUTO (parallel)')
    print(f"  {clr(C.BLUE, 'File:')} {os.path.basename(selected_file)}")
    print(f"  {clr(C.BLUE, 'Directory:')} {directory}")
    print(f"  {clr(C.BLUE, 'Mode:')} {mode_str}")
    print()

    # --- Step 2: Read phone numbers ---
    print(f"  {clr(C.BLUE, 'Reading numbers...')}")
    if selected_file.endswith('.xlsx'):
        phones = read_phones_from_xlsx(selected_file)
    else:
        # TXT/CSV fallback
        phones = []
        with open(selected_file, 'r') as f:
            for line in f:
                cleaned = re.sub(r'[^\d+]', '', line.strip())
                if cleaned and len(cleaned) >= 7:
                    if not cleaned.startswith('+'):
                        cleaned = '+' + cleaned
                    phones.append(cleaned)

    if not phones:
        print(f"  {clr(C.BRED, 'No phone numbers found!')}")
        sys.exit(1)

    print(f"  Numbers: {clr(C.BGREEN, str(len(phones)))}")
    print()

    # --- Step 3: Load accounts ---
    global accounts

    if seq_mode:
        # In seq mode, we don't need Dan.json - accounts are auto-created as needed
        raw_accounts = []
        # Still check for Dan.json in case it exists
        dan_path = None
        for candidate in ["Dan.json", "dan.json"]:
            for d in [directory, os.path.join(directory, "upload"), os.path.dirname(os.path.abspath(__file__))]:
                fp = os.path.join(d, candidate) if d else None
                if fp and os.path.exists(fp):
                    dan_path = fp
                    break
            if dan_path:
                break
        if dan_path:
            try:
                raw_accounts = load_dan_json(dan_path)
                print(f"  {clr(C.GREEN, 'Accounts loaded from Dan.json:')} {len(raw_accounts)}")
            except:
                raw_accounts = []
        if not raw_accounts and AUTO_CREATE_ACCOUNTS:
            print(f"  {clr(C.BYELLOW, 'No Dan.json — will auto-create accounts as needed')}")
    else:
        # Auto mode: find Dan.json
        dan_path = None
        for candidate in ["Dan.json", "dan.json"]:
            for d in [directory, os.path.join(directory, "upload"), os.path.dirname(os.path.abspath(__file__))]:
                fp = os.path.join(d, candidate) if d else None
                if fp and os.path.exists(fp):
                    dan_path = fp
                    break
            if dan_path:
                break

        raw_accounts = []
        if dan_path and os.path.exists(dan_path):
            try:
                raw_accounts = load_dan_json(dan_path)
                print(f"  {clr(C.GREEN, 'Accounts loaded:')} {len(raw_accounts)}")
            except Exception as e:
                print(f"  {clr(C.BYELLOW, 'Dan.json error:')} {e}")
                raw_accounts = []
        else:
            if AUTO_CREATE_ACCOUNTS:
                print(f"  {clr(C.BYELLOW, 'No Dan.json — will auto-create accounts')}")
            else:
                print(f"  {clr(C.BRED, 'No Dan.json found! Enable AUTO_CREATE_ACCOUNTS.')}")
                sys.exit(1)

    # --- Step 4: Initialize accounts ---
    print()
    print(f"  {clr(C.BCYAN, 'Initializing accounts...')}")
    accounts = init_all_accounts(raw_accounts, directory)
    if not accounts and AUTO_CREATE_ACCOUNTS:
        print(f"  {clr(C.BYELLOW, 'Auto-creating accounts...')}")
        new = auto_create_accounts(directory, ACCOUNT_CREATE_BATCH, max_retries=3)
        if new:
            accounts.extend(new)

    if not accounts and not seq_mode:
        print(f"  {clr(C.BRED, 'No accounts available!')}")
        sys.exit(1)

    mark_used.cache_dir = directory
    print(f"  {clr(C.BGREEN, 'Ready accounts:')} {len(accounts)}")
    print()

    # --- Step 5: GO! ---
    if seq_mode:
        print(f"  {clr(C.BGREEN, 'Starting concurrent sequential calling mode...')}")
        print(f"  Numbers: {len(phones)} | Workers: {num_workers} | Call duration: {SEQ_CALL_DURATION}s after answer")
        print(f"  {clr(C.DIM, f'{num_workers}x RING -> answer -> stay 62s -> hang up -> RING next -> repeat')}")
        print()
        run_sequential_mode(phones, num_workers=num_workers)
    else:
        if not accounts:
            print(f"  {clr(C.BRED, 'No accounts available!')}")
            sys.exit(1)
        print(f"  {clr(C.BGREEN, 'Starting auto calling mode...')}")
        print(f"  Numbers: {len(phones)} | Accounts: {len(accounts)} | Workers: {NUM_WORKERS}")
        print()
        run_auto_mode(phones)

    # Save cache
    print()
    print(f"  {clr(C.BLUE, 'Saving cache...')}")
    save_accounts_cache(directory, accounts, used_emails, set())
    print(f"  {clr(C.BGREEN, 'Done!')}")
    print()


if __name__ == "__main__":
    main()
